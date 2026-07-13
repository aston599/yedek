# -*- coding: utf-8 -*-
import logging
import os
import threading
import asyncio
import time
import signal
import sys
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, 
    filters, ContextTypes, CallbackQueryHandler,
    ChatMemberHandler
)
from telegram.constants import ParseMode
from config import BOT_TOKEN, SITES, APPROVED_GROUP_LINK, CHAT_GROUP_LINK, ADMIN_USER_ID, BOT_VERSION, LAST_UPDATE, AUTHORIZED_GROUP_ID, BOT_PORT, BOT_SOURCE
import database as db
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Logging ayarları
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Database güvenlik kilidi
db_lock = threading.Lock()

def parse_application_row(row):
    """
    Application row'dan verileri parse et
    Hem eski (merso) hem yeni (gameofbet) formatı destekler
    Not: Database'de kolon isimleri betpuan_username/betpuan_screenshot (geriye dönük uyumluluk)
    """
    # Row uzunluğuna göre format belirle
    if len(row) >= 13:
        # Yeni format: gameofbet kolonları var (database'de betpuan_username/betpuan_screenshot olarak)
        (app_id, user_id, username, merso_username, merso_screenshot,
         amg_username, amg_screenshot, status, application_date, approved_date,
         bot_source, gameofbet_username_db, gameofbet_screenshot_db, *rest) = row
        # Database'den gelen değerleri gameofbet olarak adlandır (ama database kolonu betpuan_username)
        betpuan_username = gameofbet_username_db
        betpuan_screenshot = gameofbet_screenshot_db
    elif len(row) >= 11:
        # Orta format: bot_source var ama gameofbet kolonları yok
        (app_id, user_id, username, merso_username, merso_screenshot,
         amg_username, amg_screenshot, status, application_date, approved_date,
         bot_source, *rest) = row
        betpuan_username = None
        betpuan_screenshot = None
    else:
        # Eski format: sadece merso kolonları var
        (app_id, user_id, username, merso_username, merso_screenshot,
         amg_username, amg_screenshot, status, application_date, approved_date,
         *rest) = row
        betpuan_username = None
        betpuan_screenshot = None
        bot_source = rest[0] if len(rest) > 0 else 'main'
    
    # Game of Bet (database'de betpuan_username olarak) varsa onu kullan, yoksa merso'yu göster
    site_username = betpuan_username if betpuan_username else merso_username
    site_screenshot = betpuan_screenshot if betpuan_screenshot else merso_screenshot
    
    return {
        'id': app_id,
        'user_id': user_id,
        'username': username,
        'site_username': site_username,  # Game of Bet veya Merso
        'site_screenshot': site_screenshot,  # Game of Bet veya Merso
        'merso_username': merso_username,
        'merso_screenshot': merso_screenshot,
        'betpuan_username': betpuan_username,
        'betpuan_screenshot': betpuan_screenshot,
        'amg_username': amg_username,
        'amg_screenshot': amg_screenshot,
        'status': status,
        'application_date': application_date,
        'approved_date': approved_date,
        'bot_source': bot_source
    }

# Rate limiting için son mesaj zamanları
user_last_message = {}
user_message_count = {}
RATE_LIMIT_SECONDS = 2  # 2 saniye içinde sadece 1 mesaj
MAX_MESSAGES_PER_MINUTE = 10  # Dakikada maksimum 10 mesaj

# Kullanıcı durumları
class UserState:
    def __init__(self):
        self.step = 0  # 0: başlangıç, 1: gameofbet screenshot, 2: gameofbet username, 3: amg screenshot, 4: amg username
        self.gameofbet_screenshot = None  # Game of Bet ekran görüntüsü (database'de betpuan_screenshot olarak kaydedilir)
        self.gameofbet_username = None  # Game of Bet kullanıcı adı (database'de betpuan_username olarak kaydedilir)
        self.amg_screenshot = None
        self.amg_username = None

user_states = {}

# Rate limiting kontrolü
def check_rate_limit(user_id):
    """Rate limit kontrolü - True dönerse mesaj gönderebilir"""
    current_time = time.time()
    
    # Son mesaj zamanını kontrol et
    if user_id in user_last_message:
        time_diff = current_time - user_last_message[user_id]
        if time_diff < RATE_LIMIT_SECONDS:
            return False, f"⏱️ Lütfen {RATE_LIMIT_SECONDS} saniye bekleyin."
    
    # Dakika başına mesaj sayısını kontrol et
    if user_id not in user_message_count:
        user_message_count[user_id] = []
    
    # 1 dakikadan eski mesajları temizle
    user_message_count[user_id] = [t for t in user_message_count[user_id] if current_time - t < 60]
    
    # Dakikada max mesaj kontrolü
    if len(user_message_count[user_id]) >= MAX_MESSAGES_PER_MINUTE:
        return False, f"⚠️ Çok fazla mesaj gönderiyorsunuz! Lütfen 1 dakika bekleyin."
    
    # Kayıt et
    user_last_message[user_id] = current_time
    user_message_count[user_id].append(current_time)
    
    return True, None

async def is_user_admin(user_id, context=None, chat_id=None):
    """Kullanıcının admin olup olmadığını kontrol et"""
    # Ana admin
    if ADMIN_USER_ID is not None and user_id == ADMIN_USER_ID:
        return True
    
    # SABİT yetkili gruptaki adminler otomatik yetki alır
    if context and AUTHORIZED_GROUP_ID:
        try:
            member = await context.bot.get_chat_member(AUTHORIZED_GROUP_ID, user_id)
            if member.status in ['creator', 'administrator']:
                return True
        except:
            pass
    
    # Veritabanındaki adminler
    return db.is_admin(user_id)

async def notify_all_admins(context, message, parse_mode=ParseMode.HTML):
    """Tüm adminlere bildirim gönder"""
    # Ana admin
    if ADMIN_USER_ID:
        try:
            await context.bot.send_message(
                chat_id=ADMIN_USER_ID,
                text=message,
                parse_mode=parse_mode
            )
        except Exception as e:
            logger.warning(f"Ana admin'e bildirim gönderilemedi: {e}")
    
    # Diğer adminler
    admins = db.get_all_admins()
    for admin in admins:
        admin_id = admin[0]
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=message,
                parse_mode=parse_mode
            )
        except Exception as e:
            logger.warning(f"Admin {admin_id}'e bildirim gönderilemedi: {e}")

async def grupid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/grupid komutu - grup ID'sini öğren (DM'den cevap)"""
    # Sadece grupta çalışır
    if update.effective_chat.type not in ['group', 'supergroup']:
        await update.message.reply_text(
            "⚠️ Bu komut sadece gruplarda çalışır!",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Sadece adminler kullanabilir (grup admini veya bot admini)
    if not await is_user_admin(update.effective_user.id, context, update.effective_chat.id):
        return
    
    chat = update.effective_chat
    group_id = chat.id
    group_title = chat.title
    group_type = chat.type
    member_count = await context.bot.get_chat_member_count(group_id)
    user_id = update.effective_user.id
    
    # Komutu sil
    try:
        await update.message.delete()
    except:
        pass
    
    # Yetkili grup mu kontrol et
    try:
        authorized_group = db.get_config('authorized_group_id')
        is_authorized = authorized_group and str(group_id) == authorized_group
    except:
        is_authorized = False
    
    message = (
        f"🆔 <b>GRUP BİLGİLERİ</b>\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📛 <b>Grup Adı:</b> {group_title}\n"
        f"🆔 <b>Grup ID:</b> <code>{group_id}</code>\n"
        f"📊 <b>Tip:</b> {group_type}\n"
        f"👥 <b>Üye Sayısı:</b> {member_count}\n\n"
    )
    
    if is_authorized:
        message += "✅ <b>Bu yetkili gruptur!</b>\n"
    else:
        message += "⚠️ <b>Bu yetkili grup değil.</b>\n"
    
    message += "\n━━━━━━━━━━━━━━━━━━━━━\n\n"
    message += "💡 <b>ID'yi kopyalamak için üzerine tıklayın.</b>"
    
    # Kullanıcıya DM'den gönder
    try:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Grup ID sorgulandı: {group_title} ({group_id}) by {user_id}")
    except Exception as e:
        logger.error(f"DM gönderme hatası: {e}")
        # DM gönderilemezse grupta geçici mesaj göster
        temp_msg = await context.bot.send_message(
            chat_id=group_id,
            text="⚠️ Size özel mesaj gönderemiyorum! Önce bana /start yazın.",
            parse_mode=ParseMode.HTML
        )
        # 5 saniye sonra sil
        await asyncio.sleep(5)
        try:
            await temp_msg.delete()
        except:
            pass

async def benkim(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanıcı ID'sini göster"""
    user_id = update.effective_user.id
    username = update.effective_user.username or "Yok"
    first_name = update.effective_user.first_name

    await update.message.reply_text(
        f"👤 **Kullanıcı Bilgileriniz:**\n\n"
        f"🆔 **User ID:** `{user_id}`\n"
        f"📝 **İsim:** {first_name}\n"
        f"🔗 **Username:** @{username}\n\n"
        f"Bu ID'yi admin olarak kaydetmek için kullanın!",
        parse_mode=ParseMode.MARKDOWN
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/yardim komutu - kullanıcı komutlarını göster"""
    
    user_help = """
📚 <b>BOT KOMUTLARI</b>

━━━━━━━━━━━━━━━━━━━━━

👤 <b>KULLANICI KOMUTLARI:</b>

🔹 <b>/start</b>
   Üyelik başvurusunu başlatır
   Site seçimi → Ekran görüntüsü → Kullanıcı adı
   
🔹 <b>/durumum</b>
   Başvuru durumunuzu gösterir
   (Bekliyor/Onaylandı/Reddedildi)
   
🔹 <b>/benkim</b>
   Telegram User ID'nizi öğrenin
   
🔹 <b>/yardim</b>
   Bu yardım mesajını gösterir

━━━━━━━━━━━━━━━━━━━━━

💡 <b>Başvuru Süreci:</b>
1. /start komutu ile başlayın
2. Game of Bet veya AMG Bahis'i seçin
3. Ekran görüntüsü yükleyin
4. Site kullanıcı adınızı yazın
5. Yönetici onayını bekleyin (5-15 dk)

━━━━━━━━━━━━━━━━━━━━━

❓ <b>Sorun mu yaşıyorsunuz?</b>
Yönetici ile iletişime geçin!
"""
    
    await update.message.reply_text(user_help, parse_mode=ParseMode.HTML)

async def admin_help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/adminyardim komutu - admin komutlarını göster"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    admin_help = """
👨‍💼 <b>ADMIN KOMUTLARI</b>

━━━━━━━━━━━━━━━━━━━━━

📋 <b>BAŞVURU YÖNETİMİ:</b>

🔸 <b>/bekleyenler [sayfa/arama]</b>
   Bekleyen başvuruları listele
   • /bekleyenler → İlk 10 başvuru
   • /bekleyenler 2 → 2. sayfa (11-20)
   • /bekleyenler kirve → Arama yap
   
🔸 <b>/incele [id]</b> veya <b>/ss [id]</b>
   Başvuru screenshot'larını gör
   • Game of Bet ve AMG screenshot'ları
   • Kullanıcı bilgileri
   
🔸 <b>/onayla [id]</b>
   Başvuruyu onayla
   • Örnek: /onayla 1
   • Kullanıcıya bildirim gönderilir
   • +18 kanal linki paylaşılır
   
🔸 <b>/reddet [id]</b>
   Başvuruyu reddet (sebep seçimi)
   • Örnek: /reddet 1
   • Red nedeni seçenekleri gösterilir
   • Kullanıcıya bildirim gönderilir
   
🔸 <b>/gecmis [sayfa/arama]</b>
   Onaylanan başvuruları göster
   • /gecmis → İlk 10 onaylı başvuru
   • /gecmis 2 → 2. sayfa
   • /gecmis merso → Arama yap

━━━━━━━━━━━━━━━━━━━━━

📊 <b>İSTATİSTİK VE RAPORLAR:</b>

🔸 <b>/istatistikler</b>
   Detaylı istatistikler
   • Toplam başvuru sayısı
   • Bekleyen/Onaylanan/Reddedilen
   • Onay/Red oranları
   • Site bazında dağılım (Game of Bet/AMG)
   • Bugünkü ve haftalık başvurular
   
🔸 <b>/exportexcel</b>
   Tüm başvuruları Excel olarak indir
   • Tüm başvuru verileri
   • Başvuru ID, User ID, Username
   • Site bilgileri, Durum, Tarihler

━━━━━━━━━━━━━━━━━━━━━

⚙️ <b>GRUP YÖNETİMİ:</b>

🔸 <b>/grupid</b>
   Grup ID'sini öğren (sadece grupta)
   • Grup bilgileri
   • Üye sayısı
   • Yetkili grup kontrolü

🔸 <b>/kirvebaslat</b>
   Grupta direkt mesaj gönder
   • Komutu otomatik siler
   • Hemen mesaj gösterir
   
🔸 <b>/kirvebasla</b>
   DM'den gruba mesaj gönder
   • Yetkili gruba gönderir
   
🔸 <b>/adminekle [user_id]</b>
   Yeni admin ekle
   • Örnek: /adminekle 123456789
   • Kullanıcı admin yetkisi alır
   
🔸 <b>/admincikar [user_id]</b>
   Admin yetkisini kaldır
   • Örnek: /admincikar 123456789
   
🔸 <b>/adminler</b>
   Tüm adminleri listele
   • Admin User ID'leri
   • Admin kullanıcı adları
   
🔸 <b>/banla [user_id] [sebep]</b>
   Kullanıcıyı engelle
   • Örnek: /banla 123456 Sahte bilgi
   • Engellenince bot kullanılamaz
   
🔸 <b>/unban [user_id]</b>
   Kullanıcının engelini kaldır
   • Örnek: /unban 123456
   
🔸 <b>/banliste</b>
   Engellenmiş kullanıcıları göster
   • Engelleme sebepleri
   • Tarihler ve detaylar
   
🔸 <b>/not [başvuru_id] [not]</b>
   Başvuruya not ekle
   • Örnek: /not 1 Kullanıcı doğrulandı
   • Admin kayıtları için
   
🔸 <b>/notlar [başvuru_id]</b>
   Başvuru notlarını göster
   • Tüm eklenen notlar
   • Tarih ve admin bilgisi
   
🔸 <b>/topluonayla [id1,id2,id3...]</b>
   Birden fazla başvuruyu onayla
   • Örnek: /topluonayla 1,2,3,4,5
   • En fazla 20 başvuru
   
🔸 <b>/toplureddet [id1,id2,id3...]</b>
   Birden fazla başvuruyu reddet
   • Örnek: /toplureddet 1,2,3
   • En fazla 20 başvuru
   
🔸 <b>/profil [user_id]</b>
   Kullanıcı profili ve geçmişi
   • Tüm başvurular
   • İstatistikler
   • Engel durumu
   
🔸 <b>!cc [sayı]</b> (Yetkili grupta)
   Mesaj silme (Şu an çalışmıyor ⚠️)
   • Bot sadece çalıştıktan sonraki mesajları silebilir
   • Botun admin olması gerekir

━━━━━━━━━━━━━━━━━━━━━

📊 <b>Admin Paneli Özellikleri:</b>
• Sayfalama: Her sayfada 10 başvuru
• Arama: ID, username, site username
• Red sebepleri: Seçenekli sistem
• Excel: Tüm veriler indirebilir
• Çoklu admin desteği
• Toplu işlemler: Max 20 başvuru
• Not sistemi: Her başvuruya not eklenebilir
• Kullanıcı profili: Detaylı geçmiş
• Engel sistemi: Kullanıcı banlama
• 24 saat spam koruması
• Screenshot doğrulama

━━━━━━━━━━━━━━━━━━━━━

🔑 <b>Admin yetkileriniz aktif!</b>
"""
    
    await update.message.reply_text(admin_help, parse_mode=ParseMode.HTML)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start komutu - sadece özel sohbette çalışır"""
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    
    # Kullanıcı engellenmiş mi kontrol et
    if db.is_user_banned(user_id):
        await update.message.reply_text(
            "🚫 <b>Engellendiniz!</b>\n\n"
            "Bu botu kullanma yetkiniz kaldırılmıştır.\n"
            "Daha fazla bilgi için yönetici ile iletişime geçin.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # 24 saat spam koruması
    last_application_time = db.get_last_application_time(user_id)
    if last_application_time:
        from datetime import datetime, timedelta
        last_time = datetime.strptime(last_application_time, '%Y-%m-%d %H:%M:%S')
        time_diff = datetime.now() - last_time
        
        if time_diff < timedelta(hours=24):
            remaining_hours = 24 - int(time_diff.total_seconds() / 3600)
            remaining_minutes = int((time_diff.total_seconds() % 3600) / 60)
            
            await update.message.reply_text(
                f"⏰ <b>Başvuru Sınırı</b>\n\n"
                f"Son başvurunuzdan sonra 24 saat geçmesi gerekiyor.\n\n"
                f"⏳ <b>Kalan Süre:</b> {remaining_hours} saat {remaining_minutes} dakika\n\n"
                f"Spam koruması için lütfen bekleyin.",
                parse_mode=ParseMode.HTML
            )
            return
    
    # Kullanıcının zaten başvurusu var mı kontrol et
    if db.check_user_has_application(user_id):
        await update.message.reply_text(
            "⚠️ **Zaten Aktif Bir Başvurunuz Var**\n\n"
            "Mevcut başvurunuz yönetici tarafından değerlendirilmeyi bekliyor.\n\n"
            "Lütfen sabırlı olun. Sonuç size bildirilecektir.",
            parse_mode=ParseMode.MARKDOWN
        )
        return
    
    # Hoş geldin mesajı
    welcome_text = """
🎯 <b>+18 Özel Kanala Hoş Geldiniz!</b>

━━━━━━━━━━━━━━━━━━━━━

📋 <b>Katılım Şartı:</b>
Özel kanalımıza katılabilmek için <b>Game of Bet VE AMG Bahis</b> sitelerinin <b>HER İKİSİNE DE</b> üye olmanız gerekmektedir.

🔐 <b>Doğrulama Süreci:</b>
1️⃣ Önce Game of Bet bilgilerini gönderin
2️⃣ Sonra AMG Bahis bilgilerini gönderin
3️⃣ Her iki site için de ekran görüntüsü + kullanıcı adı
4️⃣ Yönetici onayını bekleyin (5-15 dakika)

⚠️ <b>Önemli:</b> İki siteye de üye olmalısınız!

━━━━━━━━━━━━━━━━━━━━━

<b>Hazır mısınız? Başlayalım!</b> 👇
    """
    
    keyboard = [
        [InlineKeyboardButton("✅ Başvuruya Başla", callback_data="start_application")],
        [InlineKeyboardButton("❌ İptal", callback_data="cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        welcome_text,
        reply_markup=reply_markup,
        parse_mode=ParseMode.HTML
    )
    
    # Hoş geldin fotoğraflarını gönder
    import os
    photo_dir = "welcome_photos"
    if os.path.exists(photo_dir):
        photo_files = [f for f in os.listdir(photo_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        
        if photo_files:
            # Maksimum 3 fotoğraf gönder
            photos_to_send = photo_files[:3]
            
            # Her fotoğraf için farklı caption
            captions = [
                "🔥 Kirvem, bunlar sadece başlangıç! Başvurunu tamamla ve +18 gruba gel! 🔞",
                "💎 Özel içerikler ve sıcak sohbetler seni bekliyor! Başvurunu tamamla! 🎯",
                "🎊 VIP deneyim için hazır mısın? Başvurunu gönder ve özel gruba geçiş yap! 🚀"
            ]
            
            for i, photo_file in enumerate(photos_to_send):
                photo_path = os.path.join(photo_dir, photo_file)
                caption = captions[i] if i < len(captions) else captions[0]
                
                try:
                    with open(photo_path, 'rb') as photo:
                        await update.message.reply_photo(
                            photo=photo,
                            caption=caption,
                            parse_mode=ParseMode.HTML,
                            has_spoiler=True
                        )
                except Exception as e:
                    logger.error(f"Start fotoğraf gönderme hatası: {e}")

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Buton tıklamaları"""
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    
    if query.data == "start_application":
        # Başvuru başlat - önce Game of Bet
        if user_id not in user_states:
            user_states[user_id] = UserState()
        
        user_states[user_id].step = 1  # Game of Bet screenshot
        
        gameofbet_url = SITES["gameofbet"]["url"]
        gameofbet_name = SITES["gameofbet"]["name"]
        
        keyboard = [
            [InlineKeyboardButton(f"🌐 {gameofbet_name} Sitesine Git", url=gameofbet_url)],
            [InlineKeyboardButton("❌ İptal", callback_data="cancel")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        # Eski mesajı sil ve YENİ MESAJ gönder (görseller üstte kalsın)
        try:
            await query.message.delete()
        except:
            pass
        
        await context.bot.send_message(
            chat_id=user_id,
            text=f"<b>📍 AŞAMA 1/2: GAME OF BET</b>\n\n"
            f"<b>📝 Adım 1/2:</b> Siteye Üye Olun\n\n"
            f"🚨 <b><u>ÇOK ÖNEMLİ:</u></b> 🚨\n"
            f"Yukarıdaki butona tıklayarak <b><u>MUTLAKA BİZİM LİNKİMİZLE</u></b> kayıt olun!\n"
            f"Başka linkle kayıt olursanız <b>başvurunuz reddedilir!</b>\n\n"
            f"🎁 <b>BONUS:</b> 10.000 TL çekim imkanlı deneme bonusu!\n\n"
            f"✅ <b>{gameofbet_name}</b> sitesine gidin\n"
            f"✅ Üye olun (henüz üye değilseniz)\n"
            f"✅ Hesabınıza giriş yapın\n\n"
            f"<b>📸 Adım 2/2:</b> Ekran Görüntüsü Gönderin\n\n"
            f"Üye olduktan ve giriş yaptıktan sonra:\n"
            f"• Ana sayfanızın ekran görüntüsünü alın\n"
            f"• Kullanıcı adınız görünür olmalı\n"
            f"• Ekran görüntüsünü buraya gönderin\n\n"
            f"⬇️ <b>Game of Bet ekran görüntünüzü buraya gönderin:</b>",
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
    
    elif query.data == "cancel":
        # İptal et
        if user_id in user_states:
            del user_states[user_id]
        
        await query.edit_message_text(
            "❌ **İşlem İptal Edildi**\n\n"
            "İstediğiniz zaman /start yazarak tekrar başvurabilirsiniz.",
            parse_mode=ParseMode.MARKDOWN
        )
    
    elif query.data == "request_join":
        # Gruba katılma talebi gönder
        keyboard = [
            [InlineKeyboardButton("🔞 +18 Kanala Katılma Talebi Gönder", url=APPROVED_GROUP_LINK)]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "📲 **Gruba Katılma Talebi**\n\n"
            "Aşağıdaki butona tıklayarak özel kanalımıza katılma talebi gönderin.\n\n"
            "Talebiniz yönetici tarafından onaylandığında gruba ekleneceksiniz.",
            reply_markup=reply_markup,
            parse_mode=ParseMode.MARKDOWN
        )
    
    elif query.data.startswith("reject_"):
        # Red işlemi - format: reject_[app_id]_[reason]
        parts = query.data.split("_")
        app_id = int(parts[1])
        reason_code = parts[2]
        
        # Red nedenlerini tanımla
        reason_map = {
            "fake_ss": "Sahte ekran görüntüsü tespit edildi",
            "wrong_info": "Eksik veya hatalı bilgiler sağlandı",
            "no_membership": "Site üyeliği doğrulanamadı",
            "bad_quality": "Ekran görüntüsü kalitesi yetersiz",
            "duplicate": "Tekrarlayan başvuru",
            "no_reason": ""
        }
        
        reason = reason_map.get(reason_code, "")
        
        # Başvuruyu getir
        application = db.get_application(app_id)
        
        if not application:
            await query.answer("⚠️ Başvuru bulunamadı!", show_alert=True)
            return
        
        # Parse application data (betpuan veya merso desteği)
        app_data = parse_application_row(application)
        
        if app_data['status'] != 'pending':
            await query.answer("⚠️ Bu başvuru zaten işlenmiş!", show_alert=True)
            return
        
        # Başvuruyu reddet (sebep ile)
        db.update_application_status_with_reason(app_data['id'], 'rejected', reason)
        
        # Admin mesajını güncelle
        await query.edit_message_text(
            f"❌ <b>Başvuru Reddedildi!</b>\n\n"
            f"🆔 <b>Başvuru ID:</b> {app_data['id']}\n"
            f"👤 <b>Kullanıcı:</b> @{app_data['username']}\n"
            f"🏢 <b>Betpuan:</b> {app_data['site_username']} | <b>AMG:</b> {app_data['amg_username']}\n"
            f"❌ <b>Red Nedeni:</b> {reason if reason else 'Belirtilmedi'}\n\n"
            f"Kullanıcıya bildirim gönderildi.",
            parse_mode=ParseMode.HTML
        )
        
        # Kullanıcıya bildirim gönder
        try:
            notification = "❌ <b>Başvurunuz Reddedildi</b>\n\n"
            notification += "Üzgünüz, başvurunuz yönetici tarafından reddedildi.\n\n"
            
            if reason:
                notification += f"<b>📝 Red Nedeni:</b>\n{reason}\n\n"
            else:
                notification += "<b>📝 Red Nedeni:</b> Belirtilmedi\n\n"
            
            notification += "⏱️ <b>20 dakika sonra</b> tekrar başvuru yapabilirsiniz.\n\n"
            notification += "Bilgilerinizi düzeltip tekrar başvurmak için /start yazabilirsiniz."
            
            await context.bot.send_message(
                chat_id=user_id,
                text=notification,
                parse_mode=ParseMode.HTML
            )
            logger.info(f"Red bildirimi gönderildi: User={username}, Reason={reason}")
        except Exception as e:
            logger.error(f"Kullanıcıya bildirim gönderilemedi: {e}")
        
        await query.answer("✅ Başvuru reddedildi!")
    
    elif query.data.startswith("pending_page_"):
        # Bekleyen başvurular sayfa değiştirme
        page = int(query.data.split("_")[2])
        context.args = [str(page)]
        
        # Admin kontrolü
        if ADMIN_USER_ID and query.from_user.id == ADMIN_USER_ID:
            # Mesajı güncelle
            pending = db.get_pending_applications()
            per_page = 10
            total_pages = (len(pending) + per_page - 1) // per_page
            page = max(1, min(page, total_pages))
            
            start_idx = (page - 1) * per_page
            end_idx = min(start_idx + per_page, len(pending))
            page_pending = pending[start_idx:end_idx]
            
            message = "📋 <b>Bekleyen Başvurular:</b>\n\n"
            message += f"📄 Sayfa {page}/{total_pages} ({len(pending)} başvuru)\n\n"
            message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
            
            for app in page_pending:
                app_data = parse_application_row(app)
                
                message += (
                    f"🆔 <b>ID:</b> {app_data['id']} | 👤 @{app_data['username']}\n"
                    f"🏢 <b>Game of Bet:</b> {app_data['site_username']} | <b>AMG:</b> {app_data['amg_username']}\n"
                    f"📅 {app_data['application_date'].split('.')[0] if app_data['application_date'] else 'Bilinmiyor'}\n"
                    f"➡️ /onayla {app_data['id']} | /reddet {app_data['id']}\n\n"
                    f"━━━━━━━━━━━━━━━━━━━━━\n\n"
                )
            
            keyboard = []
            nav_buttons = []
            
            if page > 1:
                nav_buttons.append(InlineKeyboardButton("⬅️ Önceki", callback_data=f"pending_page_{page-1}"))
            if page < total_pages:
                nav_buttons.append(InlineKeyboardButton("Sonraki ➡️", callback_data=f"pending_page_{page+1}"))
            
            if nav_buttons:
                keyboard.append(nav_buttons)
            
            message += (
                "<b>💡 Kullanım:</b>\n"
                f"• /bekleyenler - Tüm bekleyenler\n"
                f"• /bekleyenler 2 - 2. sayfa\n"
                f"• /bekleyenler kirve - 'kirve' ara"
            )
            
            reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
            await query.edit_message_text(message, parse_mode=ParseMode.HTML, reply_markup=reply_markup)
        
        await query.answer()
    
    elif query.data.startswith("history_page_"):
        # Geçmiş sayfa değiştirme
        page = int(query.data.split("_")[2])
        
        # Admin kontrolü
        if ADMIN_USER_ID and query.from_user.id == ADMIN_USER_ID:
            approved = db.get_approved_applications()
            per_page = 10
            total_pages = (len(approved) + per_page - 1) // per_page
            page = max(1, min(page, total_pages))
            
            start_idx = (page - 1) * per_page
            end_idx = min(start_idx + per_page, len(approved))
            page_approved = approved[start_idx:end_idx]
            
            message = "📜 <b>Onaylanmış Başvurular:</b>\n\n"
            message += f"📄 Sayfa {page}/{total_pages} ({len(approved)} başvuru)\n\n"
            message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
            
            for app in page_approved:
                app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, app_date, approved_date, *_ = app
                
                app_date_str = app_date.split('.')[0] if app_date else "Bilinmiyor"
                approved_date_str = approved_date.split('.')[0] if approved_date else "Bilinmiyor"
                
                message += (
                    f"🆔 <b>ID:</b> {app_id} | 👤 @{username}\n"
                    f"🏢 <b>Betpuan:</b> {merso_username} | <b>AMG:</b> {amg_username}\n"
                    f"📅 <b>Başvuru:</b> {app_date_str}\n"
                    f"✅ <b>Onay:</b> {approved_date_str}\n\n"
                    f"━━━━━━━━━━━━━━━━━━━━━\n\n"
                )
            
            keyboard = []
            nav_buttons = []
            
            if page > 1:
                nav_buttons.append(InlineKeyboardButton("⬅️ Önceki", callback_data=f"history_page_{page-1}"))
            if page < total_pages:
                nav_buttons.append(InlineKeyboardButton("Sonraki ➡️", callback_data=f"history_page_{page+1}"))
            
            if nav_buttons:
                keyboard.append(nav_buttons)
            
            message += (
                f"<b>📊 Toplam:</b> {len(approved)} onaylanmış başvuru\n\n"
                "<b>💡 Kullanım:</b>\n"
                f"• /gecmis - Tüm geçmiş\n"
                f"• /gecmis 2 - 2. sayfa\n"
                f"• /gecmis kirve - 'kirve' ara"
            )
            
            reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
            await query.edit_message_text(message, parse_mode=ParseMode.HTML, reply_markup=reply_markup)
        
        await query.answer()

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fotoğraf mesajlarını işle"""
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    
    # Kullanıcının durumu var mı kontrol et
    if user_id not in user_states:
        await update.message.reply_text(
            "⚠️ Lütfen önce /start komutu ile başlayın."
        )
        return
    
    user_state = user_states[user_id]
    
    # Step 1: Betpuan screenshot || Step 3: AMG screenshot
    if user_state.step == 1 or user_state.step == 3:
        photo = update.message.photo[-1]
        
        # Dosya boyutu kontrolü (max 20MB)
        if photo.file_size and photo.file_size > 20 * 1024 * 1024:
            await update.message.reply_text(
                "⚠️ <b>Dosya Çok Büyük!</b>\n\n"
                "Lütfen 20MB'dan küçük bir görsel yükleyin.",
                parse_mode=ParseMode.HTML
            )
            return
        
        # Dosya boyutu minimum kontrolü (en az 10KB - çok küçük dosyalar şüpheli)
        if photo.file_size and photo.file_size < 10 * 1024:
            await update.message.reply_text(
                "⚠️ <b>Dosya Çok Küçük!</b>\n\n"
                "Lütfen geçerli bir ekran görüntüsü yükleyin.",
                parse_mode=ParseMode.HTML
            )
            return
        
        file = await context.bot.get_file(photo.file_id)
        
        # Screenshots klasörünü oluştur
        os.makedirs("screenshots", exist_ok=True)
        
        # Game of Bet için step 1, AMG için step 3
        site_name = "gameofbet" if user_state.step == 1 else "amg"
        screenshot_path = f"screenshots/{user_id}_{site_name}.jpg"
        await file.download_to_drive(screenshot_path)
        
        # Dosya formatı kontrolü (gerçek bir resim mi?)
        try:
            from PIL import Image
            img = Image.open(screenshot_path)
            
            # Minimum boyut kontrolü (en az 200x200 pixel)
            if img.width < 200 or img.height < 200:
                os.remove(screenshot_path)
                await update.message.reply_text(
                    "⚠️ <b>Görsel Çok Küçük!</b>\n\n"
                    "Lütfen en az 200x200 boyutunda bir ekran görüntüsü yükleyin.",
                    parse_mode=ParseMode.HTML
                )
                return
            
            # Format kontrolü
            if img.format not in ['JPEG', 'PNG', 'WEBP']:
                os.remove(screenshot_path)
                await update.message.reply_text(
                    "⚠️ <b>Geçersiz Format!</b>\n\n"
                    "Lütfen JPEG, PNG veya WEBP formatında görsel yükleyin.",
                    parse_mode=ParseMode.HTML
                )
                return
            
            img.close()
            logger.info(f"Screenshot doğrulandı: {screenshot_path} ({img.width}x{img.height}, {img.format})")
        except Exception as e:
            if os.path.exists(screenshot_path):
                os.remove(screenshot_path)
            logger.error(f"Screenshot doğrulama hatası: {e}")
            await update.message.reply_text(
                "⚠️ <b>Geçersiz Dosya!</b>\n\n"
                "Lütfen geçerli bir görsel dosyası yükleyin.",
                parse_mode=ParseMode.HTML
            )
            return
        
        # Game of Bet için step 1 -> 2, AMG için step 3 -> 4
        if user_state.step == 1:
            user_state.gameofbet_screenshot = screenshot_path
            user_state.step = 2
            site_display = SITES["gameofbet"]["name"]
        else:
            user_state.amg_screenshot = screenshot_path
            user_state.step = 4
            site_display = SITES["amg"]["name"]
        
        await update.message.reply_text(
            f"✅ <b>Ekran Görüntüsü Alındı!</b>\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📝 <b>Şimdi Kullanıcı Adı</b>\n\n"
            f"Lütfen <b>{site_display}</b> sitesindeki kullanıcı adınızı yazın:\n\n"
            f"💡 <b>Örnek:</b> kullanici123\n\n"
            f"⬇️ <b>Kullanıcı adınızı buraya yazın:</b>",
            parse_mode=ParseMode.HTML
        )
    else:
        await update.message.reply_text(
            "⚠️ Şu anda ekran görüntüsü beklemiyorum. Lütfen adımları takip edin."
        )

async def handle_application_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Başvuru sürecindeki metin mesajlarını işle"""
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    username = update.effective_user.username or update.effective_user.first_name or f"User{user_id}"
    
    # Kullanıcının durumu var mı kontrol et
    if user_id not in user_states:
        return
    
    user_state = user_states[user_id]
    
    try:
        # Step 2: Game of Bet username || Step 4: AMG username
        if user_state.step == 2:
            # Game of Bet kullanıcı adı alındı
            user_state.gameofbet_username = update.message.text.strip()
            user_state.step = 3  # AMG screenshot bekleniyor
            
            amg_url = SITES["amg"]["url"]
            amg_name = SITES["amg"]["name"]
            
            keyboard = [
                [InlineKeyboardButton(f"🌐 {amg_name} Sitesine Git", url=amg_url)]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"✅ <b>Game of Bet Tamamlandı!</b>\n\n"
                f"<b>📍 AŞAMA 2/2: AMG BAHIS</b>\n\n"
                f"<b>📝 Adım 1/2:</b> Siteye Üye Olun\n\n"
                f"🚨 <b><u>ÇOK ÖNEMLİ:</u></b> 🚨\n"
                f"Yukarıdaki butona tıklayarak <b><u>MUTLAKA BİZİM LİNKİMİZLE</u></b> kayıt olun!\n"
                f"Başka linkle kayıt olursanız <b>başvurunuz reddedilir!</b>\n\n"
                f"✅ <b>{amg_name}</b> sitesine gidin\n"
                f"✅ Üye olun (henüz üye değilseniz)\n"
                f"✅ Hesabınıza giriş yapın\n\n"
                f"<b>📸 Adım 2/2:</b> Ekran Görüntüsü Gönderin\n\n"
                f"⬇️ <b>AMG Bahis ekran görüntünüzü buraya gönderin:</b>",
                reply_markup=reply_markup,
                parse_mode=ParseMode.HTML
            )
        
        elif user_state.step == 4:
            # AMG kullanıcı adı alındı - başvuru tamamlandı!
            user_state.amg_username = update.message.text.strip()
            
            # Başvuruyu kaydet (HER İKİ SİTE İÇİN) - Thread safe
            with db_lock:
                application_id = db.add_application(
                    user_id=user_id,
                    username=username,
                    merso_username=None,  # Eski kolon - boş bırakıyoruz (ileride kullanılabilir)
                    merso_screenshot=None,  # Eski kolon - boş bırakıyoruz
                    amg_username=user_state.amg_username,
                    amg_screenshot=user_state.amg_screenshot,
                    betpuan_username=user_state.gameofbet_username,  # Game of Bet verisi (database'de betpuan_username kolonunda)
                    betpuan_screenshot=user_state.gameofbet_screenshot,  # Game of Bet verisi (database'de betpuan_screenshot kolonunda)
                    bot_source=BOT_SOURCE
                )
            
            if application_id:
                logger.info(f"Yeni başvuru: ID={application_id}, User={username}, Game of Bet={user_state.gameofbet_username}, AMG={user_state.amg_username}")
                
                # Gerçek zamanlı kullanıcı bilgilerini al
                user_info = await get_user_info(context, user_id)
                
                # Kullanıcı adını oluştur
                if user_info['full_name'] != 'Bilinmiyor':
                    display_name = user_info['full_name']
                    if user_info['username']:
                        display_name += f" (@{user_info['username']})"
                elif user_info['username']:
                    display_name = f"@{user_info['username']}"
                elif username:
                    display_name = f"@{username}"
                else:
                    display_name = f"ID: {user_id}"
                
                # Adminlere bildirim gönder - KISALTILMIŞ
                admin_notification = (
                    f"🔔 <b>YENİ BAŞVURU!</b>\n\n"
                    f"🆔 Başvuru ID: <code>{application_id}</code>\n"
                    f"🤖 Bot Kaynağı: <b>{BOT_SOURCE.upper()}</b>\n"
                    f"👤 Kullanıcı: {display_name}\n"
                    f"🆔 User ID: <code>{user_id}</code>\n\n"
                    f"🏢 <b>GAME OF BET:</b>\n"
                    f"📝 Kullanıcı: {user_state.gameofbet_username}\n\n"
                    f"🎯 <b>AMG BAHIS:</b>\n"
                    f"📝 Kullanıcı: {user_state.amg_username}\n\n"
                    f"Başvuruyu incelemek için:\n"
                    f"/incele {application_id}\n\n"
                    f"Hızlı işlem:\n"
                    f"/onayla {application_id}\n"
                    f"/reddet {application_id}"
                )
                await notify_all_admins(context, admin_notification)
                
                # Başarı mesajı + grup linkleri
                keyboard = [
                    [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                await update.message.reply_text(
                   "✅ <b>Başvuru Alındı!</b>\n\n"
                   f"<b>Bilgileriniz:</b>\n"
                   f"• Game of Bet: {user_state.gameofbet_username}\n"
                   f"• AMG: {user_state.amg_username}\n\n"
                   f"<b>Ne Olacak?</b>\n"
                   f"• Yönetici inceleyecek\n"
                   f"• 5-15 dk içinde cevap\n"
                   f"• Bildirim gelecek\n\n"
                   f"<b>Şimdi ne yapmalı?</b>\n"
                   f"Aşağıdaki butonlara tıkla ve bekle!\n\n"
                   f"Onay sonrası otomatik ekleneceksin! 🎉",
                   reply_markup=reply_markup,
                   parse_mode=ParseMode.HTML
               )
                
                # Kullanıcı durumunu temizle
                del user_states[user_id]
            else:
                await update.message.reply_text(
                    "❌ <b>Bir Hata Oluştu</b>\n\n"
                    "Başvurunuz kaydedilemedi. Lütfen tekrar deneyin veya yönetici ile iletişime geçin.",
                    parse_mode=ParseMode.HTML
                )
        else:
            # Yanlış adımda metin gönderildi
            await update.message.reply_text(
                "⚠️ <b>Yanlış Adım!</b>\n\n"
                "Lütfen adımları sırayla takip edin. Baştan başlamak için /start yazın.",
                parse_mode=ParseMode.HTML
            )
    
    except Exception as e:
        logger.error(f"handle_application_text hatası (User: {user_id}, Step: {user_state.step}): {e}", exc_info=True)
        
        # Kullanıcıya anlaşılır hata mesajı gönder
        await update.message.reply_text(
            "⚠️ <b>Bir hata oluştu!</b>\n\n"
            "Lütfen tekrar deneyin veya /start yazarak baştan başlayın.\n\n"
            "Sorun devam ederse yönetici ile iletişime geçin.",
            parse_mode=ParseMode.HTML
        )
        
        # Adminlere hata bildirimi gönder
        error_notification = (
            f"⚠️ <b>BAŞVURU HATASI!</b>\n\n"
            f"👤 Kullanıcı: @{username} ({user_id})\n"
            f"📝 Adım: {user_state.step}\n"
            f"❌ Hata: <code>{str(e)[:200]}</code>\n\n"
            f"Lütfen kontrol edin!"
        )
        try:
            await notify_all_admins(context, error_notification)
        except:
            pass

async def handle_general_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Genel sohbet mesajlarını işle"""
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    
    # Rate limit kontrolü
    can_send, error_msg = check_rate_limit(user_id)
    if not can_send:
        await update.message.reply_text(error_msg, parse_mode=ParseMode.HTML)
        return
    
    message_text = update.message.text.lower()
    
    # Selamlaşma
    if any(word in message_text for word in ['selam', 'merhaba', 'hi', 'hello', 'hey']):
        await update.message.reply_text(
            "👋 <b>Merhaba!</b> Hoş geldiniz!\n\n"
            "🔞 +18 kanalımıza katılmak için /start yazın!\n"
            "❓ Yardım için /yardim yazın!",
            parse_mode=ParseMode.HTML
        )
    
    # Teşekkür
    elif any(word in message_text for word in ['teşekkür', 'sağol', 'thanks', 'teşekkürler', 'sağolun']):
        await update.message.reply_text(
            "😊 <b>Rica ederim!</b>\n\n"
            "Başka bir konuda yardımcı olabilir miyim?",
            parse_mode=ParseMode.HTML
        )
    
    # Nasılsın
    elif any(word in message_text for word in ['nasılsın', 'nasıl', 'how are', 'nasılsın bot']):
        await update.message.reply_text(
            "🤖 <b>Ben iyiyim, teşekkürler!</b>\n\n"
            "Size nasıl yardımcı olabilirim?",
            parse_mode=ParseMode.HTML
        )
    
    # Durum sorgulama
    elif any(word in message_text for word in ['durum', 'durumum', 'ne oldu', 'başvuru']):
        if db.check_user_has_application(user_id):
            await update.message.reply_text(
                "📋 <b>Başvurunuz mevcut!</b>\n\n"
                "✅ Durum: Değerlendirme bekliyor\n"
                "⏱️ Süre: 5-15 dakika\n\n"
                "📊 Detay için: /durumum",
                parse_mode=ParseMode.HTML
            )
        else:
            await update.message.reply_text(
                "🎯 <b>Henüz başvurunuz yok!</b>\n\n"
                "📝 <b>Gereksinimler:</b>\n"
                "• Game of Bet üyeliği\n"
                "• AMG Bahis üyeliği\n\n"
                "🚀 <b>Başlamak için:</b> /start",
                parse_mode=ParseMode.HTML
            )
    
    # Yardım isteği
    elif any(word in message_text for word in ['yardım', 'help', 'nasıl', 'ne yapmalı', 'bilmiyorum']):
        await update.message.reply_text(
            "❓ <b>Yardım Menüsü</b>\n\n"
            "📋 <b>Mevcut komutlar:</b>\n"
            "• /start - Başvuru yap\n"
            "• /durumum - Başvuru durumu\n"
            "• /profil - Profil bilgisi\n"
            "• /yardim - Detaylı yardım\n\n"
            "💡 <b>İpucu:</b> /start yazarak başvuruya başlayabilirsiniz!",
            parse_mode=ParseMode.HTML
        )
    
    # Bilinmeyen mesaj
    else:
        await update.message.reply_text(
            "🤔 <b>Anlayamadım!</b>\n\n"
            "📋 <b>Mevcut komutlar:</b>\n"
            "• /start - Başvuru yap\n"
            "• /durumum - Başvuru durumu\n"
            "• /yardim - Yardım menüsü\n\n"
            "💡 <b>İpucu:</b> /start yazarak başvuruya başlayabilirsiniz!",
            parse_mode=ParseMode.HTML
        )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ana metin mesaj handler - öncelik sırası ile"""
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    
    try:
        # 1️⃣ ÖNCE: Başvuru sürecinde mi kontrol et
        if user_id in user_states:
            # Başvuru sürecindeki kullanıcı - mevcut sistem
            await handle_application_text(update, context)
            return
        
        # 2️⃣ SONRA: Genel sohbet
        await handle_general_chat(update, context)
        
    except Exception as e:
        logger.error(f"Text handler error: {e}")
        await update.message.reply_text(
            "⚠️ <b>Bir hata oluştu!</b>\n\n"
            "Lütfen tekrar deneyin veya /start yazarak baştan başlayın.",
            parse_mode=ParseMode.HTML
        )

async def delete_service_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Grup servis mesajlarını otomatik sil (join/leave) - TÜM GRUPLARDA"""
    # Grup değilse çık
    if update.effective_chat.type not in ['group', 'supergroup']:
        return
    
    # Yeni üyeler varsa onlara hoş geldin mesajı gönder
    if update.message and update.message.new_chat_members:
        for new_member in update.message.new_chat_members:
            # Bot kendisini atlasın
            if new_member.is_bot:
                continue
            
            # Yeni üyeye hoş geldin mesajı gönder (DM engelliyse hata veriyor, kapatıldı)
            # await send_welcome_to_new_member(context, new_member)
    
    # Servis mesajı mı kontrol et (katılma/ayrılma mesajları)
    if update.message and (
        update.message.new_chat_members or 
        update.message.left_chat_member
    ):
        try:
            await update.message.delete()
            logger.info(f"Servis mesajı silindi: Grup={update.effective_chat.title}, Mesaj ID={update.message.message_id}")
        except Exception as e:
            logger.warning(f"Servis mesajı silinemedi: {e}")

async def test_welcome_photos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/testfoto komutu - hoş geldin mesajını ve fotoğrafları test et"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    user_id = update.effective_user.id
    first_name = update.effective_user.first_name
    
    await update.message.reply_text("📸 <b>Test başlatılıyor...</b>", parse_mode=ParseMode.HTML)
    
    # Hoş geldin mesajı
    welcome_text = (
        f"🔥 <b>Kirvem {first_name}, Hoş Geldin!</b>\n\n"
        f"Deniz Aksoy'un özel +18 grubuna katıldın ama tam erişim için bazı şartları yerine getirmelisin.\n\n"
        f"<b>Gereksinimler:</b>\n"
        f"• Game of Bet üyeliği\n"
        f"• AMG Bahis üyeliği\n\n"
        f"<b>Süreç çok basit:</b>\n"
        f"1. Aşağıdaki butona bas\n"
        f"2. Screenshot gönder\n"
        f"3. 5-15 dk'da onay al\n\n"
        f"Onay aldıktan sonra +18 kanala erişim sağlayacaksın! 🎉\n\n"
        f"<b>Hemen başla! 👇</b>"
    )
    
    # Buton
    keyboard = [
        [InlineKeyboardButton("✅ Başvuru Yap", url=f"https://t.me/{context.bot.username}?start=verify")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Mesajı gönder
    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_text,
        parse_mode=ParseMode.HTML,
        reply_markup=reply_markup
    )
    
    # Spoilerli resimleri gönder
    welcome_photos_dir = "welcome_photos"
    photo_count = 0
    
    if os.path.exists(welcome_photos_dir):
        photo_files = [f for f in os.listdir(welcome_photos_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))]
        
        # Farklı captionlar - merak uyandırıcı
        captions = [
            "🔥 <b>Kirvem, sadece küçük bir önizleme...</b>\n\n"
            "Gerçek şov +18 kanalında başlıyor! 🔞\n\n"
            "Başvurunu tamamla, seni bekliyoruz! 💋",
            
            "😈 <b>Bu sadece ısınma turuydu...</b>\n\n"
            "Asıl eğlence +18 grubunda seni bekliyor! 🔥\n\n"
            "Hemen başvuru yap, pişman olmazsın! 😏",
            
            "🎁 <b>Daha neler var neler...</b>\n\n"
            "Bunlar sadece başlangıç kirve! Devamı efsane! 🔞\n\n"
            "Başvurunu yap, aramıza katıl! 🔥"
        ]
        
        for idx, photo_file in enumerate(photo_files[:3]):  # En fazla 3 resim
            photo_path = os.path.join(welcome_photos_dir, photo_file)
            caption = captions[idx] if idx < len(captions) else captions[0]
            
            try:
                await context.bot.send_photo(
                    chat_id=user_id,
                    photo=open(photo_path, 'rb'),
                    caption=caption,
                    parse_mode=ParseMode.HTML,
                    has_spoiler=True  # Spoiler yapıyor!
                )
                photo_count += 1
            except Exception as e:
                await update.message.reply_text(f"❌ Resim gönderilemedi {photo_file}: {e}")
                logger.error(f"Test: Resim gönderilemedi {photo_file}: {e}")
    else:
        await update.message.reply_text("⚠️ welcome_photos klasörü bulunamadı!")
    
    # Sonuç
    await update.message.reply_text(
        f"✅ <b>Test tamamlandı!</b>\n\n"
        f"📸 Gönderilen resim: {photo_count}\n"
        f"📁 Klasör: {welcome_photos_dir}",
        parse_mode=ParseMode.HTML
    )
    
    logger.info(f"Hoş geldin test edildi by {update.effective_user.id}")

async def send_welcome_to_new_member(context, new_member):
    """Yeni üyeye özel mesajdan hoş geldin mesajı gönder"""
    try:
        user_id = new_member.id
        first_name = new_member.first_name
        username = new_member.username
        
        # Hoş geldin mesajı
        welcome_text = (
            f"🔥 <b>Kirvem {first_name}, Hoş Geldin!</b>\n\n"
            f"Deniz Aksoy'un özel +18 grubuna katıldın ama tam erişim için bazı şartları yerine getirmelisin.\n\n"
            f"<b>Gereksinimler:</b>\n"
            f"• Game of Bet üyeliği\n"
            f"• AMG Bahis üyeliği\n\n"
            f"<b>Süreç çok basit:</b>\n"
            f"1. Aşağıdaki butona bas\n"
            f"2. Screenshot gönder\n"
            f"3. 5-15 dk'da onay al\n\n"
            f"Onay aldıktan sonra +18 kanala erişim sağlayacaksın! 🎉\n\n"
            f"<b>Hemen başla! 👇</b>"
        )
        
        # Buton
        keyboard = [
            [InlineKeyboardButton("✅ Başvuru Yap", url=f"https://t.me/{context.bot.username}?start=verify")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        # Mesajı gönder
        await context.bot.send_message(
            chat_id=user_id,
            text=welcome_text,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        # Spoilerli resimleri gönder
        welcome_photos_dir = "welcome_photos"
        if os.path.exists(welcome_photos_dir):
            photo_files = [f for f in os.listdir(welcome_photos_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))]
            
            # Farklı captionlar - merak uyandırıcı
            captions = [
                "🔥 <b>Kirvem, sadece küçük bir önizleme...</b>\n\n"
                "Gerçek şov +18 kanalında başlıyor! 🔞\n\n"
                "Başvurunu tamamla, seni bekliyoruz! 💋",
                
                "😈 <b>Bu sadece ısınma turuydu...</b>\n\n"
                "Asıl eğlence +18 grubunda seni bekliyor! 🔥\n\n"
                "Hemen başvuru yap, pişman olmazsın! 😏",
                
                "🎁 <b>Daha neler var neler...</b>\n\n"
                "Bunlar sadece başlangıç kirve! Devamı efsane! 🔞\n\n"
                "Başvurunu yap, aramıza katıl! 🔥"
            ]
            
            for idx, photo_file in enumerate(photo_files[:3]):  # En fazla 3 resim
                photo_path = os.path.join(welcome_photos_dir, photo_file)
                caption = captions[idx] if idx < len(captions) else captions[0]
                
                try:
                    await context.bot.send_photo(
                        chat_id=user_id,
                        photo=open(photo_path, 'rb'),
                        caption=caption,
                        parse_mode=ParseMode.HTML,
                        has_spoiler=True  # Spoiler yapıyor!
                    )
                except Exception as e:
                    logger.error(f"Resim gönderilemedi {photo_file}: {e}")
        
        logger.info(f"Hoş geldin mesajı gönderildi: User={username} ({user_id})")
        
    except Exception as e:
        logger.error(f"Hoş geldin mesajı gönderilemedi: {e}")
        # Kullanıcı botu bloklamış olabilir, sessizce geç

async def my_chat_member_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bot gruba/kanala eklendiğinde - OTOMATİK YETKİLENDİRME"""
    chat_member = update.my_chat_member
    
    # Bot admin yapıldığında veya üye olduğunda
    if chat_member.new_chat_member.status in ['administrator', 'member']:
        chat = update.effective_chat
        user = update.my_chat_member.from_user
        
        # Grup, süper grup veya kanal
        if chat.type in ['group', 'supergroup', 'channel']:
            chat_id = chat.id
            chat_title = chat.title
            chat_type = "Kanal" if chat.type == 'channel' else "Grup"
            
            # GÜVENLİK: Sadece bot sahibi ekleyebilir
            if user.id != ADMIN_USER_ID:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=f"🚫 <b>YETKİSİZ ERİŞİM!</b>\n\n"
                         f"Sadece bot sahibi beni ekleyebilir.",
                    parse_mode=ParseMode.HTML
                )
                logger.warning(f"Yetkisiz ekleme denemesi: {user.username} ({user.id}) - {chat_title}")
                await context.bot.leave_chat(chat_id)
                return
            
            # Bot admin mi kontrol et
            try:
                bot_member = await context.bot.get_chat_member(chat_id, context.bot.id)
                is_admin = bot_member.status == 'administrator'
            except:
                is_admin = False
            
            if not is_admin and chat.type != 'channel':
                await context.bot.send_message(
                    chat_id=chat_id,
                    text="⚠️ <b>Yetersiz Yetki!</b>\n\n"
                         "Beni <b>ADMIN</b> olarak ekleyin.",
                    parse_mode=ParseMode.HTML
                )
                return
            
            # Grubu/Kanalı kaydet
            try:
                db.add_main_group(chat_id, chat_title)
                
                # Eğer bu ana yetkili grup ise işaretle
                if chat_id == AUTHORIZED_GROUP_ID:
                    db.set_config('authorized_group_id', str(chat_id))
                
                logger.info(f"{chat_type} kaydedildi: {chat_title} ({chat_id})")
                
                # Bot sahibine ÖZEL MESAJ gönder
                notification_text = f"""
✅ <b>BOT KURULDU!</b>

{chat_type}: <b>{chat_title}</b>
ID: <code>{chat_id}</code>

━━━━━━━━━━━━━━━━━━━━━

📝 <b>Admin Komutları:</b>
/grupid - Bu {chat_type.lower()}un ID'sini göster
/bekleyenler - Başvuruları gör
/onayla [id] - Onayla
/reddet [id] - Reddet
/kirvebaslat - Duyuru gönder (grupta)
/adminyardim - Tüm komutlar

🚀 <b>Bot hazır ve aktif!</b>
                """
                
                # Bot sahibine özel mesaj
                try:
                    await context.bot.send_message(
                        chat_id=ADMIN_USER_ID,
                        text=notification_text,
                        parse_mode=ParseMode.HTML
                    )
                except Exception as e:
                    logger.error(f"Admin bildirimi gönderilemedi: {e}")
            except Exception as e:
                logger.error(f"{chat_type} kaydetme hatası: {e}")
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=f"⚠️ <b>Kurulum Hatası!</b>\n\n{str(e)}",
                    parse_mode=ParseMode.HTML
                )

async def get_user_info(context, user_id):
    """Kullanıcı bilgilerini Telegram'dan gerçek zamanlı al"""
    try:
        user = await context.bot.get_chat(user_id)
        return {
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'full_name': f"{user.first_name or ''} {user.last_name or ''}".strip()
        }
    except Exception as e:
        logger.warning(f"Kullanıcı bilgisi alınamadı {user_id}: {e}")
        return {
            'username': None,
            'first_name': None,
            'last_name': None,
            'full_name': 'Bilinmiyor'
        }

async def admin_pending_applications(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/bekleyenler [sayfa] [arama] komutu - bekleyen başvuruları göster"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Sayfa numarası ve arama parametresi
    page = 1
    search_query = None
    
    if context.args:
        try:
            page = int(context.args[0])
        except ValueError:
            # İlk parametre sayı değilse, arama terimi olarak kabul et
            search_query = ' '.join(context.args).lower()
            page = 1
        
        # Eğer ikinci parametre varsa, o arama terimidir
        if len(context.args) > 1:
            try:
                int(context.args[0])  # İlk parametre sayı mı kontrol et
                search_query = ' '.join(context.args[1:]).lower()
            except ValueError:
                pass
    
    # Bekleyen başvuruları getir
    all_pending = db.get_pending_applications()
    
    # Arama filtresi uygula
    if search_query:
        filtered = []
        for app in all_pending:
            app_data = parse_application_row(app)
            # ID, username veya site username'de ara
            if (search_query in str(app_data['id']) or 
                search_query in (app_data['username'] or '').lower() or 
                search_query in (app_data['site_username'] or '').lower() or
                search_query in (app_data['amg_username'] or '').lower()):
                filtered.append(app)
        pending = filtered
    else:
        pending = all_pending
    
    if not pending:
        if search_query:
            await update.message.reply_text(
                f"🔍 <b>Arama Sonucu</b>\n\n"
                f"'{search_query}' için başvuru bulunamadı.",
                parse_mode=ParseMode.HTML
            )
        else:
            await update.message.reply_text(
                "✅ <b>Bekleyen Başvuru Yok</b>\n\n"
                "Şu anda onay bekleyen başvuru bulunmuyor.",
                parse_mode=ParseMode.HTML
            )
        return
    
    # Sayfalama
    per_page = 10
    total_pages = (len(pending) + per_page - 1) // per_page
    page = max(1, min(page, total_pages))
    
    start_idx = (page - 1) * per_page
    end_idx = min(start_idx + per_page, len(pending))
    page_pending = pending[start_idx:end_idx]
    
    # Başvuruları listele
    message = "📋 <b>Bekleyen Başvurular:</b>\n\n"
    if search_query:
        message += f"🔍 Arama: '{search_query}'\n\n"
    message += f"📄 Sayfa {page}/{total_pages} ({len(pending)} başvuru)\n\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    for app in page_pending:
        app_data = parse_application_row(app)
        
        # Gerçek zamanlı kullanıcı bilgilerini al
        user_info = await get_user_info(context, app_data['user_id'])
        
        # Kullanıcı adını oluştur
        if user_info['full_name'] != 'Bilinmiyor':
            display_name = user_info['full_name']
            if user_info['username']:
                display_name += f" (@{user_info['username']})"
        elif user_info['username']:
            display_name = f"@{user_info['username']}"
        elif app_data['username']:
            display_name = f"@{app_data['username']}"
        else:
            display_name = f"ID: {app_data['user_id']}"
        
        message += (
            f"📌 <b>BAŞVURU #{app_data['id']}</b>\n"
            f"👤 Kullanıcı: {display_name}\n"
            f"🆔 ID: <code>{app_data['user_id']}</code>\n"
            f"🏢 Game of Bet: <code>{app_data['site_username']}</code>\n"
            f"🎯 AMG: <code>{app_data['amg_username']}</code>\n"
            f"📅 {app_data['application_date'].split('.')[0] if app_data['application_date'] else 'Bilinmiyor'}\n\n"
            f"🔍 <b>İNCELE:</b> /incele {app_data['id']}\n"
            f"✅ <b>ONAYLA:</b> /onayla {app_data['id']}\n"
            f"❌ <b>REDDET:</b> /reddet {app_data['id']}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n\n"
        )
    
    # Navigasyon butonları
    keyboard = []
    nav_buttons = []
    
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Önceki", callback_data=f"pending_page_{page-1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Sonraki ➡️", callback_data=f"pending_page_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Kullanım bilgisi
    message += (
        "<b>💡 Kullanım:</b>\n"
        f"• /bekleyenler - Tüm bekleyenler\n"
        f"• /bekleyenler 2 - 2. sayfa\n"
        f"• /bekleyenler kirve - 'kirve' ara\n"
        f"• /bekleyenler 2 kirve - 2. sayfa, 'kirve' ara"
    )
    
    reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
    await update.message.reply_text(message, parse_mode=ParseMode.HTML, reply_markup=reply_markup)

async def admin_view_screenshots(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/ss [id] veya /incele [id] komutu - başvuru screenshot'larını gör"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Başvuru ID'sini al
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /ss [başvuru_id]")
        return
    
    try:
        app_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz başvuru ID'si!")
        return
    
    # Başvuruyu getir
    application = db.get_application(app_id)
    
    if not application:
        await update.message.reply_text("⚠️ Başvuru bulunamadı!")
        return
    
    app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, application_date, approved_date, *_ = application
    
    # Gerçek zamanlı kullanıcı bilgilerini al
    user_info = await get_user_info(context, user_id)
    
    # Kullanıcı adını oluştur
    if user_info['full_name'] != 'Bilinmiyor':
        display_name = user_info['full_name']
        if user_info['username']:
            display_name += f" (@{user_info['username']})"
    elif user_info['username']:
        display_name = f"@{user_info['username']}"
    elif username:
        display_name = f"@{username}"
    else:
        display_name = f"ID: {user_id}"
    
    # Başvuru bilgileri
    info_message = (
        f"📸 <b>BAŞVURU SCREENSHOT'LARI</b>\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🆔 <b>Başvuru ID:</b> {app_id}\n"
        f"👤 <b>Kullanıcı:</b> {display_name}\n"
        f"🆔 <b>User ID:</b> <code>{user_id}</code>\n"
        f"📊 <b>Durum:</b> {status}\n\n"
        f"🏢 <b>Merso:</b> {merso_username}\n"
        f"🎯 <b>AMG:</b> {amg_username}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📸 Screenshot'lar aşağıda:\n"
    )
    
    await update.message.reply_text(info_message, parse_mode=ParseMode.HTML)
    
    # Merso Screenshot
    if merso_screenshot and os.path.exists(merso_screenshot):
        try:
            await update.message.reply_photo(
                photo=open(merso_screenshot, 'rb'),
                caption=f"🏢 <b>GAME OF BET</b>\n\n"
                        f"👤 Kullanıcı: {merso_username}",
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"Merso screenshot gönderme hatası: {e}")
            await update.message.reply_text(f"❌ Merso screenshot gönderilemedi: {e}")
    else:
        await update.message.reply_text("⚠️ Merso screenshot bulunamadı!")
    
    # AMG Screenshot
    if amg_screenshot and os.path.exists(amg_screenshot):
        try:
            await update.message.reply_photo(
                photo=open(amg_screenshot, 'rb'),
                caption=f"🎯 <b>AMG BAHIS</b>\n\n"
                        f"👤 Kullanıcı: {amg_username}",
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"AMG screenshot gönderme hatası: {e}")
            await update.message.reply_text(f"❌ AMG screenshot gönderilemedi: {e}")
    else:
        await update.message.reply_text("⚠️ AMG screenshot bulunamadı!")
    
    logger.info(f"Screenshot görüntülendi: App={app_id} by {update.effective_user.id}")

async def admin_approve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/onayla [id] komutu"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Başvuru ID'sini al
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /onayla [başvuru_id]")
        return
    
    try:
        app_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz başvuru ID'si!")
        return
    
    # Başvuruyu getir
    application = db.get_application(app_id)
    
    if not application:
        await update.message.reply_text("⚠️ Başvuru bulunamadı!")
        return
    
    app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, date, *_ = application
    
    if status != 'pending':
        await update.message.reply_text("⚠️ Bu başvuru zaten işlenmiş!")
        return
    
    # Başvuruyu onayla
    db.update_application_status(app_id, 'approved')
    
    # Gerçek zamanlı kullanıcı bilgilerini al
    user_info = await get_user_info(context, user_id)
    
    # Kullanıcı adını oluştur
    if user_info['full_name'] != 'Bilinmiyor':
        display_name = user_info['full_name']
        if user_info['username']:
            display_name += f" (@{user_info['username']})"
    elif user_info['username']:
        display_name = f"@{user_info['username']}"
    elif username:
        display_name = f"@{username}"
    else:
        display_name = f"ID: {user_id}"
    
    # Admin mesajı
    await update.message.reply_text(
        f"✅ <b>Başvuru Onaylandı!</b>\n\n"
        f"🆔 Başvuru ID: {app_id}\n"
        f"👤 Kullanıcı: {display_name}\n"
        f"🆔 User ID: <code>{user_id}</code>\n"
        f"🏢 Merso: {merso_username}\n"
        f"🎯 AMG: {amg_username}",
        parse_mode=ParseMode.HTML
    )
    
    # Kullanıcıya bildirim gönder
    try:
        keyboard = [
            [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                "🎉 <b>Tebrikler Kirvem!</b>\n\n"
                "Başvurun onaylandı! Artık aramızdasın.\n\n"
                "Aşağıdaki butona tıklayarak +18 kanalına katılabilirsin.\n\n"
                "Eğlenceli sohbetler seni bekliyor! 🔥"
            ),
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Onay bildirimi gönderildi: User={username}")
    except Exception as e:
        logger.error(f"Kullanıcıya bildirim gönderilemedi: {e}")

async def admin_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/reddet [id] [sebep] komutu - manuel sebep yazabilir veya butonlardan seçebilir"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Başvuru ID'sini al
    if not context.args:
        await update.message.reply_text(
            "⚠️ <b>Kullanım:</b>\n\n"
            "📝 <b>Manuel sebep:</b> /reddet [id] [sebep]\n"
            "Örnek: /reddet 8 Sahte bilgi kullandınız\n\n"
            "🔘 <b>Buton ile:</b> /reddet [id]\n"
            "Örnek: /reddet 8",
            parse_mode=ParseMode.HTML
        )
        return
    
    try:
        app_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz başvuru ID'si!")
        return
    
    # Başvuruyu getir
    application = db.get_application(app_id)
    
    if not application:
        await update.message.reply_text("⚠️ Başvuru bulunamadı!")
        return
    
    app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, application_date, approved_date, *_ = application
    
    if status != 'pending':
        await update.message.reply_text("⚠️ Bu başvuru zaten işlenmiş!")
        return
    
    # Eğer sebep yazılmışsa direkt reddet
    if len(context.args) > 1:
        custom_reason = " ".join(context.args[1:])
        
        # Başvuruyu reddet
        db.update_application_status_with_reason(app_id, 'rejected', custom_reason)
        
        # Gerçek zamanlı kullanıcı bilgilerini al
        user_info = await get_user_info(context, user_id)
        
        # Kullanıcı adını oluştur
        if user_info['full_name'] != 'Bilinmiyor':
            display_name = user_info['full_name']
            if user_info['username']:
                display_name += f" (@{user_info['username']})"
        elif user_info['username']:
            display_name = f"@{user_info['username']}"
        elif username:
            display_name = f"@{username}"
        else:
            display_name = f"ID: {user_id}"
        
        # Admin'e onay mesajı
        await update.message.reply_text(
            f"❌ <b>Başvuru Reddedildi!</b>\n\n"
            f"🆔 <b>Başvuru ID:</b> {app_id}\n"
            f"👤 <b>Kullanıcı:</b> {display_name}\n"
            f"🆔 <b>User ID:</b> <code>{user_id}</code>\n"
            f"🏢 <b>Merso:</b> {merso_username} | <b>AMG:</b> {amg_username}\n"
            f"❌ <b>Red Nedeni:</b> {custom_reason}\n\n"
            f"Kullanıcıya bildirim gönderildi.",
            parse_mode=ParseMode.HTML
        )
        
        # Kullanıcıya bildirim gönder
        try:
            notification = "❌ <b>Başvurunuz Reddedildi</b>\n\n"
            notification += "Üzgünüz, başvurunuz yönetici tarafından reddedildi.\n\n"
            notification += f"<b>📝 Red Nedeni:</b>\n{custom_reason}\n\n"
            notification += "⏱️ <b>20 dakika sonra</b> tekrar başvuru yapabilirsiniz.\n\n"
            notification += "Bilgilerinizi düzeltip tekrar başvurmak için /start yazabilirsiniz."
            
            await context.bot.send_message(
                chat_id=user_id,
                text=notification,
                parse_mode=ParseMode.HTML
            )
            logger.info(f"Red bildirimi gönderildi: User={username}, Reason={custom_reason}")
        except Exception as e:
            logger.error(f"Kullanıcıya bildirim gönderilemedi: {e}")
        
        return
    
    # Sebep yazılmamışsa butonları göster
    keyboard = [
        [InlineKeyboardButton("❌ Sahte Ekran Görüntüsü", callback_data=f"reject_{app_id}_fake_ss")],
        [InlineKeyboardButton("⚠️ Eksik/Hatalı Bilgiler", callback_data=f"reject_{app_id}_wrong_info")],
        [InlineKeyboardButton("🚫 Üyelik Doğrulanamadı", callback_data=f"reject_{app_id}_no_membership")],
        [InlineKeyboardButton("📸 Kalitesiz Ekran Görüntüsü", callback_data=f"reject_{app_id}_bad_quality")],
        [InlineKeyboardButton("🔄 Tekrarlayan Başvuru", callback_data=f"reject_{app_id}_duplicate")],
        [InlineKeyboardButton("❌ Red Et (Neden Belirtmeden)", callback_data=f"reject_{app_id}_no_reason")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Gerçek zamanlı kullanıcı bilgilerini al
    user_info = await get_user_info(context, user_id)
    
    # Kullanıcı adını oluştur
    if user_info['full_name'] != 'Bilinmiyor':
        display_name = user_info['full_name']
        if user_info['username']:
            display_name += f" (@{user_info['username']})"
    elif user_info['username']:
        display_name = f"@{user_info['username']}"
    elif username:
        display_name = f"@{username}"
    else:
        display_name = f"ID: {user_id}"
    
    await update.message.reply_text(
        f"❓ <b>Red Nedeni Seçin</b>\n\n"
        f"🆔 <b>Başvuru ID:</b> {app_id}\n"
        f"👤 <b>Kullanıcı:</b> {display_name}\n"
        f"🆔 <b>User ID:</b> <code>{user_id}</code>\n"
        f"🏢 <b>Merso:</b> {merso_username} | <b>AMG:</b> {amg_username}\n\n"
        f"💡 <b>İpucu:</b> Manuel sebep yazmak için:\n"
        f"<code>/reddet {app_id} [sebep]</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_markup
    )

async def admin_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/gecmis [sayfa] [arama] komutu - onaylanan başvuruları göster"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Sayfa numarası ve arama parametresi
    page = 1
    search_query = None
    
    if context.args:
        try:
            page = int(context.args[0])
        except ValueError:
            # İlk parametre sayı değilse, arama terimi olarak kabul et
            search_query = ' '.join(context.args).lower()
            page = 1
        
        # Eğer ikinci parametre varsa, o arama terimidir
        if len(context.args) > 1:
            try:
                int(context.args[0])  # İlk parametre sayı mı kontrol et
                search_query = ' '.join(context.args[1:]).lower()
            except ValueError:
                pass
    
    # Onaylanan başvuruları getir
    all_approved = db.get_approved_applications()
    
    # Arama filtresi uygula
    if search_query:
        filtered = []
        for app in all_approved:
            app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, app_date, approved_date, *_ = app
            # ID, username veya site username'de ara
            if (search_query in str(app_id) or 
                search_query in username.lower() or 
                search_query in (merso_username or "").lower() or
                search_query in (amg_username or "").lower()):
                filtered.append(app)
        approved = filtered
    else:
        approved = all_approved
    
    if not approved:
        if search_query:
            await update.message.reply_text(
                f"🔍 <b>Arama Sonucu</b>\n\n"
                f"'{search_query}' için onaylanmış başvuru bulunamadı.",
                parse_mode=ParseMode.HTML
            )
        else:
            await update.message.reply_text(
                "✅ <b>Onaylanmış Başvuru Yok</b>\n\n"
                "Henüz onaylanmış başvuru bulunmuyor.",
                parse_mode=ParseMode.HTML
            )
        return
    
    # Sayfalama
    per_page = 10
    total_pages = (len(approved) + per_page - 1) // per_page
    page = max(1, min(page, total_pages))
    
    start_idx = (page - 1) * per_page
    end_idx = min(start_idx + per_page, len(approved))
    page_approved = approved[start_idx:end_idx]
    
    # Başvuruları listele
    message = "📜 <b>Onaylanmış Başvurular:</b>\n\n"
    if search_query:
        message += f"🔍 Arama: '{search_query}'\n\n"
    message += f"📄 Sayfa {page}/{total_pages} ({len(approved)} başvuru)\n\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    for app in page_approved:
        app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, app_date, approved_date, *_ = app
        
        # Gerçek zamanlı kullanıcı bilgilerini al
        user_info = await get_user_info(context, user_id)
        
        # Kullanıcı adını oluştur
        if user_info['full_name'] != 'Bilinmiyor':
            display_name = user_info['full_name']
            if user_info['username']:
                display_name += f" (@{user_info['username']})"
        elif user_info['username']:
            display_name = f"@{user_info['username']}"
        elif username:
            display_name = f"@{username}"
        else:
            display_name = f"ID: {user_id}"
        
        # Tarihleri düzenle
        app_date_str = app_date.split('.')[0] if app_date else "Bilinmiyor"
        approved_date_str = approved_date.split('.')[0] if approved_date else "Bilinmiyor"
        
        message += (
            f"🆔 <b>ID:</b> {app_id} | 👤 {display_name}\n"
            f"🆔 <b>User ID:</b> <code>{user_id}</code>\n"
            f"🏢 <b>Merso:</b> {merso_username} | <b>AMG:</b> {amg_username}\n"
            f"📅 <b>Başvuru:</b> {app_date_str}\n"
            f"✅ <b>Onay:</b> {approved_date_str}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n\n"
        )
    
    # Navigasyon butonları
    keyboard = []
    nav_buttons = []
    
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Önceki", callback_data=f"history_page_{page-1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Sonraki ➡️", callback_data=f"history_page_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Kullanım bilgisi
    message += (
        f"<b>📊 Toplam:</b> {len(approved)} onaylanmış başvuru\n\n"
        "<b>💡 Kullanım:</b>\n"
        f"• /gecmis - Tüm geçmiş\n"
        f"• /gecmis 2 - 2. sayfa\n"
        f"• /gecmis kirve - 'kirve' ara\n"
        f"• /gecmis 2 kirve - 2. sayfa, 'kirve' ara"
    )
    
    reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
    await update.message.reply_text(message, parse_mode=ParseMode.HTML, reply_markup=reply_markup)

async def user_application_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/durumum komutu - kullanıcının başvuru durumunu göster"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    
    # Kullanıcının başvurusunu getir
    application = db.get_user_application(user_id)
    
    if not application:
        await update.message.reply_text(
            "❓ <b>Başvuru Bulunamadı</b>\n\n"
            "Henüz bir başvurunuz bulunmuyor.\n\n"
            "Başvuru yapmak için /start komutunu kullanın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, app_date, approved_date, *_ = application
    
    # Durum emojileri
    status_emoji = {
        'pending': '⏳',
        'approved': '✅',
        'rejected': '❌'
    }
    
    status_text = {
        'pending': 'Beklemede',
        'approved': 'Onaylandı',
        'rejected': 'Reddedildi'
    }
    
    message = f"{status_emoji.get(status, '📋')} <b>Başvuru Durumunuz</b>\n\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
    message += f"🆔 <b>Başvuru ID:</b> {app_id}\n"
    message += f"📊 <b>Durum:</b> {status_text.get(status, status)}\n"
    message += f"🏢 <b>Merso:</b> {merso_username} | <b>AMG:</b> {amg_username}\n"
    message += f"📅 <b>Başvuru Tarihi:</b> {app_date.split('.')[0] if app_date else 'Bilinmiyor'}\n"
    
    if status == 'approved' and approved_date:
        message += f"✅ <b>Onay Tarihi:</b> {approved_date.split('.')[0]}\n"
    elif status == 'rejected':
        # Red nedenini kontrol et
        try:
            cursor = db.sqlite3.connect(db.DATABASE_PATH).cursor()
            cursor.execute('SELECT red_reason FROM applications WHERE id = ?', (app_id,))
            result = cursor.fetchone()
            if result and result[0]:
                message += f"\n❌ <b>Red Nedeni:</b> {result[0]}\n"
        except:
            pass
    
    message += "\n━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    if status == 'pending':
        message += "⏳ <b>Başvurunuz inceleniyor...</b>\n"
        message += "Yönetici onayı bekleniyor. Lütfen sabırlı olun."
    elif status == 'approved':
        keyboard = [
            [InlineKeyboardButton("🔞 +18 Kanala Git", url=APPROVED_GROUP_LINK)]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        message += "🎉 <b>Tebrikler! Başvurunuz onaylandı.</b>\n"
        message += "Aşağıdaki butona tıklayarak kanala katılabilirsiniz."
        await update.message.reply_text(message, parse_mode=ParseMode.HTML, reply_markup=reply_markup)
        return
    elif status == 'rejected':
        message += "😞 <b>Başvurunuz reddedildi.</b>\n\n"
        message += "⏱️ <b>20 dakika sonra</b> tekrar başvuru yapabilirsiniz.\n"
        message += "Tekrar başvurmak için /start komutunu kullanabilirsiniz."
    
    await update.message.reply_text(message, parse_mode=ParseMode.HTML)

async def reset_my_application(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/sifirla komutu - kullanıcının başvurusunu sıfırla"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    user_id = update.effective_user.id
    
    # Aktif başvuru durumunu temizle
    if user_id in user_states:
        del user_states[user_id]
    
    # Database'den başvuruyu sil
    application = db.get_user_application(user_id)
    
    if not application:
        await update.message.reply_text(
            "ℹ️ <b>Başvuru Bulunamadı</b>\n\n"
            "Zaten bir başvurunuz bulunmuyor.\n\n"
            "Yeni başvuru yapmak için /start komutunu kullanın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    app_id = application[0]
    
    # Başvuruyu sil
    try:
        conn = db.sqlite3.connect(db.DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM applications WHERE user_id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            "✅ <b>Başvuru Sıfırlandı!</b>\n\n"
            f"🗑️ Başvuru ID <code>{app_id}</code> silindi.\n\n"
            "Yeni başvuru yapmak için /start komutunu kullanabilirsiniz.",
            parse_mode=ParseMode.HTML
        )
        
        logger.info(f"Başvuru sıfırlandı: User={user_id}, App ID={app_id}")
        
    except Exception as e:
        logger.error(f"Başvuru sıfırlama hatası: {e}")
        await update.message.reply_text(
            "❌ <b>Hata!</b>\n\n"
            "Başvuru sıfırlanırken bir hata oluştu. Lütfen tekrar deneyin.",
            parse_mode=ParseMode.HTML
        )

async def admin_reset_application(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/basvurusil [user_id] komutu - Admin: kullanıcının başvurusunu sil"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "⚠️ <b>Kullanım:</b> /basvurusil [user_id]\n\n"
            "<b>Örnek:</b> /basvurusil 123456789",
            parse_mode=ParseMode.HTML
        )
        return
    
    try:
        target_user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz User ID!")
        return
    
    # Kullanıcının başvurusunu getir
    application = db.get_user_application(target_user_id)
    
    if not application:
        await update.message.reply_text(
            f"ℹ️ <b>Başvuru Bulunamadı</b>\n\n"
            f"🆔 User ID: <code>{target_user_id}</code>\n\n"
            f"Bu kullanıcının başvurusu bulunmuyor.",
            parse_mode=ParseMode.HTML
        )
        return
    
    app_id = application[0]
    username = application[2] or "Bilinmiyor"
    
    # Başvuruyu sil
    try:
        # Aktif durum varsa temizle
        if target_user_id in user_states:
            del user_states[target_user_id]
        
        # Database'den sil
        conn = db.sqlite3.connect(db.DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM applications WHERE user_id = ?', (target_user_id,))
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            "✅ <b>Başvuru Silindi!</b>\n\n"
            f"👤 Kullanıcı: {username}\n"
            f"🆔 User ID: <code>{target_user_id}</code>\n"
            f"🗑️ Başvuru ID: <code>{app_id}</code>\n\n"
            "Başvuru başarıyla silindi.",
            parse_mode=ParseMode.HTML
        )
        
        logger.info(f"[ADMIN] Başvuru silindi: User={target_user_id}, App ID={app_id} by {update.effective_user.id}")
        
        # Kullanıcıya bildirim gönder
        try:
            await context.bot.send_message(
                chat_id=target_user_id,
                text="ℹ️ <b>Başvurunuz Silindi</b>\n\n"
                     "Yönetici tarafından başvurunuz silindi.\n\n"
                     "Yeni başvuru yapmak için /start komutunu kullanabilirsiniz.",
                parse_mode=ParseMode.HTML
            )
        except:
            pass
        
    except Exception as e:
        logger.error(f"Başvuru silme hatası: {e}")
        await update.message.reply_text(
            "❌ <b>Hata!</b>\n\n"
            "Başvuru silinirken bir hata oluştu. Lütfen tekrar deneyin.",
            parse_mode=ParseMode.HTML
        )

async def admin_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/istatistikler komutu - başvuru istatistiklerini göster"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # İstatistikleri getir
    stats = db.get_statistics()
    
    message = "📊 <b>Başvuru İstatistikleri</b>\n\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    # Genel istatistikler
    message += "<b>📈 Genel Durum:</b>\n"
    message += f"  • Toplam Başvuru: <b>{stats['total']}</b>\n"
    message += f"  • ⏳ Bekleyen: <b>{stats['pending']}</b>\n"
    message += f"  • ✅ Onaylanan: <b>{stats['approved']}</b>\n"
    message += f"  • ❌ Reddedilen: <b>{stats['rejected']}</b>\n\n"
    
    # Onay oranı
    if stats['total'] > 0:
        approval_rate = (stats['approved'] / stats['total']) * 100
        rejection_rate = (stats['rejected'] / stats['total']) * 100
        message += f"<b>📊 Oranlar:</b>\n"
        message += f"  • Onay Oranı: <b>{approval_rate:.1f}%</b>\n"
        message += f"  • Red Oranı: <b>{rejection_rate:.1f}%</b>\n\n"
    
    # Site bazında istatistikler
    if stats['by_site']:
        message += "<b>🏢 Site Bazında:</b>\n"
        for site_key, count in stats['by_site'].items():
            site_display = SITES.get(site_key, {}).get('name', site_key)
            message += f"  • {site_display}: <b>{count}</b>\n"
        message += "\n"
    
    # Zaman bazlı istatistikler
    message += "<b>📅 Zaman Bazlı:</b>\n"
    message += f"  • Bugün: <b>{stats['today']}</b>\n"
    message += f"  • Son 7 Gün: <b>{stats['week']}</b>\n\n"
    
    message += "━━━━━━━━━━━━━━━━━━━━━\n\n"
    message += f"<i>Son güncelleme: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
    
    await update.message.reply_text(message, parse_mode=ParseMode.HTML)

async def admin_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/exportexcel komutu - tüm başvuruları Excel'e aktar"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    await update.message.reply_text("📊 Excel dosyası hazırlanıyor...")
    
    try:
        # Tüm başvuruları getir
        applications = db.get_all_applications()
        
        if not applications:
            await update.message.reply_text("❌ Henüz başvuru bulunmuyor.")
            return
        
        # Excel workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Basvurular"
        
        # Başlıkları ekle
        headers = ["ID", "Kullanıcı ID", "Kullanıcı Adı", "Site Kullanıcı Adları (Merso/AMG)", 
                   "Durum", "Başvuru Tarihi", "Onay Tarihi"]
        ws.append(headers)
        
        # Başlık stilini ayarla
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Verileri ekle
        for app in applications:
            app_id, user_id, username, merso_username, merso_screenshot, amg_username, amg_screenshot, status, app_date, approved_date, *_ = app
            
            status_tr = {'pending': 'Bekliyor', 'approved': 'Onaylandı', 'rejected': 'Reddedildi'}.get(status, status)
            
            ws.append([
                app_id,
                user_id,
                username or "Yok",
                f"{merso_username} / {amg_username}",
                status_tr,
                app_date or "",
                approved_date or ""
            ])
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        # Dosyayı kaydet
        filename = f"basvurular_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        # Dosyayı gönder
        with open(filename, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=filename,
                caption=f"📊 <b>Başvuru Raporu</b>\n\n"
                        f"Toplam: {len(applications)} başvuru\n"
                        f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                parse_mode=ParseMode.HTML
            )
        
        # Dosyayı sil
        os.remove(filename)
        
    except Exception as e:
        logger.error(f"Excel export hatası: {e}")
        await update.message.reply_text(f"❌ Excel dosyası oluşturulurken hata oluştu: {e}")

async def admin_send_welcome_in_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kirvebaslat komutu - GRUPTA direkt mesaj gönder"""
    # Sadece grupta çalışır
    if update.effective_chat.type not in ['group', 'supergroup']:
        return
    
    # Sadece adminler kullanabilir (grup admini veya bot admini)
    if not await is_user_admin(update.effective_user.id, context, update.effective_chat.id):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Bu grubu yetkili grup olarak kaydet (yoksa)
    group_id = update.effective_chat.id
    try:
        authorized_group = db.get_config('authorized_group_id')
        if not authorized_group:
            db.set_config('authorized_group_id', str(group_id))
            logger.info(f"Yetkili grup ayarlandı: {group_id}")
    except:
        pass
    
    # Hoş geldin mesajı - GÜÇLÜ VE ÇEKİCİ
    welcome_message = """
🔥 <b>VIP +18 ÖZEL GRUBA HOŞ GELDİNİZ!</b> 🔥

💎 <b>VIP +18 Özel Dünya</b>
Sınırsız eğlence, özel içerikler ve premium ayrıcalıklar seni bekliyor!

━━━━━━━━━━━━━━━━━━━━
<b>📌 GEREKSİNİMLER:</b>
✅ Game of Bet Premium Üyelik
✅ AMG Bahis Premium Üyelik

<b>⚡️ HIZLI ONAY SÜRECİ:</b>
1️⃣ Üyelik Onaylat butonuna tıkla
2️⃣ Her iki siteden screenshot gönder
3️⃣ 5-15 dakika içinde VIP erişim!

━━━━━━━━━━━━━━━━━━━━
🎁 <b>SADECESİNE ÖZEL AVANTAJLAR:</b>
• Premium +18 içerikler
• Özel kapalı grup erişimi
• Sınırsız eğlence garantisi
• 7/24 aktif topluluk

<b>🚀 HEMEN BAŞLA VE ÖZEL DÜNYAMIZA KATIL! 👇</b>
    """
    
    # Butonlar
    keyboard = [
        [InlineKeyboardButton("✅ Üyelik Onaylat", url=f"https://t.me/{context.bot.username}?start=verify")],
        [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)],
        [InlineKeyboardButton("💬 Sohbet Grubuna Katıl", url=CHAT_GROUP_LINK)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        # Komutu sil
        try:
            await update.message.delete()
        except:
            pass
        
        # Bu gruba direkt mesaj gönder
        await context.bot.send_message(
            chat_id=group_id,
            text=welcome_message,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        logger.info(f"Hoş geldin mesajı gönderildi: Group={group_id}")
        
    except Exception as e:
        logger.error(f"Mesaj gönderme hatası: {e}")
        await update.message.reply_text(f"❌ Mesaj gönderilemedi: {e}")

async def admin_send_welcome_theme2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kirvebasla2 komutu - tema 2: VIP odaklı"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Yetkili grubu al
    try:
        authorized_group = db.get_config('authorized_group_id')
    except:
        authorized_group = None
    
    if not authorized_group:
        await update.message.reply_text(
            "⚠️ <b>Henüz yetkili grup ayarlanmamış!</b>\n\n"
            "Grupta <code>/kirvebaslat</code> yazın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # TEMA 2: VIP Odaklı
    welcome_message = """
👑 <b>VIP KIRVE AİLESİNE HOŞ GELDİNİZ!</b>

🎯 <b>Özel Davet:</b>
Deniz Aksoy'un seçkin +18 VIP grubuna katılım için özel şartlar bulunmaktadır.

💎 <b>VIP Üyelik Şartları:</b>
• Game of Bet VIP üyeliği
• AMG Bahis VIP üyeliği

⚡ <b>Hızlı Onay Süreci:</b>
1️⃣ VIP başvuru butonuna bas
2️⃣ Doğrulama screenshot'larını gönder  
3️⃣ 5-15 dakikada VIP onayı al

🚀 <b>VIP'e geçiş için hazır mısın?</b>
    """
    
    keyboard = [
        [InlineKeyboardButton("✅ Üyelik Onaylat", url=f"https://t.me/{context.bot.username}?start=verify")],
        [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)],
        [InlineKeyboardButton("💬 Sohbet Grubuna Katıl", url=CHAT_GROUP_LINK)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await context.bot.send_message(
            chat_id=int(authorized_group),
            text=welcome_message,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        logger.info(f"VIP tema mesajı gönderildi: Group={authorized_group}")
        
    except Exception as e:
        logger.error(f"Mesaj gönderme hatası: {e}")
        await update.message.reply_text(f"❌ Mesaj gönderilemedi: {e}")

async def admin_send_welcome_theme3(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kirvebasla3 komutu - tema 3: Eğlence odaklı"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Yetkili grubu al
    try:
        authorized_group = db.get_config('authorized_group_id')
    except:
        authorized_group = None
    
    if not authorized_group:
        await update.message.reply_text(
            "⚠️ <b>Henüz yetkili grup ayarlanmamış!</b>\n\n"
            "Grupta <code>/kirvebaslat</code> yazın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # TEMA 3: Eğlence Odaklı
    welcome_message = """
🎉 <b>EĞLENCE BAŞLIYOR!</b>

🔥 <b>Deniz Aksoy'un Özel Partisine Katıl!</b>
Sıcak sohbetler, özel içerikler ve sınırsız eğlence seni bekliyor!

🎯 <b>Katılım Koşulları:</b>
• Game of Bet üyeliği
• AMG Bahis üyeliği

⚡ <b>Hızlı Giriş:</b>
1. Başvuru butonuna bas
2. Screenshot'ları gönder
3. 5-15 dk'da partiye katıl!

🎊 <b>Eğlenceye hazır mısın?</b>
    """
    
    keyboard = [
        [InlineKeyboardButton("✅ Üyelik Onaylat", url=f"https://t.me/{context.bot.username}?start=verify")],
        [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)],
        [InlineKeyboardButton("💬 Sohbet Grubuna Katıl", url=CHAT_GROUP_LINK)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await context.bot.send_message(
            chat_id=int(authorized_group),
            text=welcome_message,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        logger.info(f"Eğlence tema mesajı gönderildi: Group={authorized_group}")
        
    except Exception as e:
        logger.error(f"Mesaj gönderme hatası: {e}")
        await update.message.reply_text(f"❌ Mesaj gönderilemedi: {e}")

async def admin_send_welcome_theme4(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kirvebasla4 komutu - tema 4: Profesyonel"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Yetkili grubu al
    try:
        authorized_group = db.get_config('authorized_group_id')
    except:
        authorized_group = None
    
    if not authorized_group:
        await update.message.reply_text(
            "⚠️ <b>Henüz yetkili grup ayarlanmamış!</b>\n\n"
            "Grupta <code>/kirvebaslat</code> yazın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # TEMA 4: Profesyonel
    welcome_message = """
📋 <b>KIRVEHUB ÜYELİK BAŞVURUSU</b>

🏢 <b>Deniz Aksoy Özel Grubu</b>
Kaliteli içerik ve özel üye deneyimi için tasarlanmıştır.

📝 <b>Başvuru Gereksinimleri:</b>
• Game of Bet platform üyeliği
• AMG Bahis platform üyeliği

⏱️ <b>Onay Süreci:</b>
1. Başvuru formunu doldurun
2. Doğrulama belgelerini yükleyin
3. 5-15 dakika içinde sonuç alın

✅ <b>Başvurunuzu tamamlayın:</b>
    """
    
    keyboard = [
        [InlineKeyboardButton("✅ Üyelik Onaylat", url=f"https://t.me/{context.bot.username}?start=verify")],
        [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)],
        [InlineKeyboardButton("💬 Sohbet Grubuna Katıl", url=CHAT_GROUP_LINK)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await context.bot.send_message(
            chat_id=int(authorized_group),
            text=welcome_message,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        logger.info(f"Profesyonel tema mesajı gönderildi: Group={authorized_group}")
        
    except Exception as e:
        logger.error(f"Mesaj gönderme hatası: {e}")
        await update.message.reply_text(f"❌ Mesaj gönderilemedi: {e}")

async def admin_send_welcome_theme5(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kirvebasla2 komutu - tema 5: Çoklu Başvuru Sistemi"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Yetkili grubu al
    try:
        authorized_group = db.get_config('authorized_group_id')
    except:
        authorized_group = None
    
    if not authorized_group:
        await update.message.reply_text(
            "⚠️ <b>Henüz yetkili grup ayarlanmamış!</b>\n\n"
            "Grupta <code>/kirvebaslat</code> yazın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # TEMA 5: Çoklu Başvuru Sistemi
    welcome_message = """
🎯 <b>ÇOKLU BAŞVURU SİSTEMİ</b>

🔥 <b>Deniz Aksoy Özel Grubu</b>
Artık istediğiniz türde başvuru yapabilirsiniz!

📋 <b>Başvuru Türleri:</b>

🔵 <b>Tam Başvuru:</b> Game of Bet + AMG (Her ikisi de gerekli)
🟢 <b>Sadece AMG:</b> Sadece AMG Bahis başvurusu  
🟡 <b>Sadece Game of Bet:</b> Sadece Game of Bet başvurusu

⚡ <b>Hızlı Onay:</b> 5-15 dakika
🎯 <b>Esnek Seçim:</b> İstediğiniz türü seçin

🚨 <u><b>ÖNEMLİ:</b></u> 🚨
<u><b>MUTLAKA BİZİM LİNKİMİZLE</b></u> kayıt olun!
    """
    
    keyboard = [
        [InlineKeyboardButton("🔵 Tam Başvuru (Merso + AMG)", url=f"https://t.me/{context.bot.username}?start=verify")],
        [InlineKeyboardButton("🟢 Sadece AMG Başvuru", url="https://t.me/KirveBot2Bot?start=verify")],
        [InlineKeyboardButton("🟡 Sadece Merso Başvuru", url="https://t.me/KirveBot2Bot?start=verify")],
        [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)],
        [InlineKeyboardButton("💬 Sohbet Grubuna Katıl", url=CHAT_GROUP_LINK)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await context.bot.send_message(
            chat_id=int(authorized_group),
            text=welcome_message,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        logger.info(f"Çoklu başvuru tema mesajı gönderildi: Group={authorized_group}")
        
    except Exception as e:
        logger.error(f"Mesaj gönderme hatası: {e}")
        await update.message.reply_text(f"❌ Mesaj gönderilemedi: {e}")

async def admin_send_welcome(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kirvebasla komutu - özel mesajdan yetkili gruba mesaj gönder"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Yetkili grubu al
    try:
        authorized_group = db.get_config('authorized_group_id')
    except:
        authorized_group = None
    
    if not authorized_group:
        await update.message.reply_text(
            "⚠️ <b>Henüz yetkili grup ayarlanmamış!</b>\n\n"
            "Grupta <code>/kirvebaslat</code> yazın.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Hoş geldin mesajı - GÜÇLÜ VE ÇEKİCİ
    welcome_message = """
🔥 <b>VIP +18 ÖZEL GRUBA HOŞ GELDİNİZ!</b> 🔥

💎 <b>VIP +18 Özel Dünya</b>
Sınırsız eğlence, özel içerikler ve premium ayrıcalıklar seni bekliyor!

━━━━━━━━━━━━━━━━━━━━
<b>📌 GEREKSİNİMLER:</b>
✅ Game of Bet Premium Üyelik
✅ AMG Bahis Premium Üyelik

<b>⚡️ HIZLI ONAY SÜRECİ:</b>
1️⃣ Üyelik Onaylat butonuna tıkla
2️⃣ Her iki siteden screenshot gönder
3️⃣ 5-15 dakika içinde VIP erişim!

━━━━━━━━━━━━━━━━━━━━
🎁 <b>SADECESİNE ÖZEL AVANTAJLAR:</b>
• Premium +18 içerikler
• Özel kapalı grup erişimi
• Sınırsız eğlence garantisi
• 7/24 aktif topluluk

<b>🚀 HEMEN BAŞLA VE ÖZEL DÜNYAMIZA KATIL! 👇</b>
    """
    
    keyboard = [
        [InlineKeyboardButton("✅ Üyelik Onaylat", url=f"https://t.me/{context.bot.username}?start=verify")],
        [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)],
        [InlineKeyboardButton("💬 Sohbet Grubuna Katıl", url=CHAT_GROUP_LINK)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await context.bot.send_message(
            chat_id=int(authorized_group),
            text=welcome_message,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        
        await update.message.reply_text(
            "✅ <b>Mesaj gönderildi!</b>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {e}")

async def admin_clear_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """!cc [sayı] komutu - belirlenen sayıda mesajı sil"""
    # Sadece grupta çalışır
    if update.effective_chat.type not in ['group', 'supergroup']:
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        return
    
    # Yetkili grupta mı kontrol et
    authorized_group = db.get_config('authorized_group_id')
    if not authorized_group or str(update.effective_chat.id) != authorized_group:
        return
    
    # Mesaj sayısını al (MessageHandler ile args çalışmaz, manuel parse edelim)
    try:
        # Mesaj metninden sayıyı çıkar: "!cc 50" -> 50
        import re
        match = re.match(r'^!cc(?:\s+(\d+))?$', update.message.text)
        if match and match.group(1):
            delete_count = int(match.group(1))
        else:
            delete_count = 1  # Varsayılan: sadece !cc mesajını sil
        
        # Maksimum 100 mesaj sınırı
        delete_count = min(delete_count, 100)
        
        logger.info(f"!cc komutu alındı: {delete_count} mesaj silinecek")
        
        # Önce komut mesajını sil
        try:
            await update.message.delete()
        except:
            pass
        
        # Geriye doğru mesajları sil
        deleted = 0
        current_message_id = update.message.message_id - 1  # Komut mesajından bir önceki
        
        # delete_count kadar mesaj silmeye çalış
        failed = 0
        for i in range(delete_count):
            if current_message_id > 0:
                try:
                    await context.bot.delete_message(
                        chat_id=update.effective_chat.id,
                        message_id=current_message_id
                    )
                    deleted += 1
                    logger.info(f"Mesaj silindi: ID={current_message_id}")
                except Exception as e:
                    # Mesaj bulunamazsa veya silinemezse devam et
                    failed += 1
                    logger.warning(f"Mesaj silinemedi (ID: {current_message_id}): {e}")
                current_message_id -= 1
            else:
                break
        
        logger.info(f"Silme özeti: Toplam {delete_count} denendi, {deleted} silindi, {failed} başarısız")
        
        # Bilgilendirme mesajı gönder ve 3 saniye sonra sil
        info_msg = await update.effective_chat.send_message(
            f"🗑️ {deleted} mesaj silindi.",
            parse_mode=ParseMode.HTML
        )
        
        # 3 saniye bekle ve bilgilendirme mesajını da sil
        import asyncio
        await asyncio.sleep(3)
        try:
            await info_msg.delete()
        except:
            pass
        
        logger.info(f"Mesajlar silindi: {deleted} adet, User={update.effective_user.username}")
        
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz sayı! Kullanım: !cc [sayı]")
    except Exception as e:
        logger.error(f"Mesaj silme hatası: {e}")

async def admin_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/adminekle [user_id] komutu - yeni admin ekle"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece ana admin kullanabilir
    if ADMIN_USER_ID is None or update.effective_user.id != ADMIN_USER_ID:
        await update.message.reply_text("⚠️ Bu komutu sadece ana yönetici kullanabilir.")
        return
    
    # User ID kontrolü
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /adminekle [user_id]")
        return
    
    try:
        new_admin_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz User ID!")
        return
    
    # Kendini eklemeye çalışıyor mu?
    if new_admin_id == ADMIN_USER_ID:
        await update.message.reply_text("⚠️ Siz zaten ana adminsiniz!")
        return
    
    # Admin ekle
    success = db.add_admin(new_admin_id, None, ADMIN_USER_ID)
    
    if success:
        await update.message.reply_text(
            f"✅ <b>Admin Eklendi!</b>\n\n"
            f"🆔 User ID: <code>{new_admin_id}</code>\n"
            f"🔑 Admin yetkileri verildi.\n\n"
            f"Kullanıcı artık tüm admin komutlarını kullanabilir!",
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Yeni admin eklendi: {new_admin_id} by {ADMIN_USER_ID}")
    else:
        await update.message.reply_text(
            f"⚠️ <b>Bu kullanıcı zaten admin!</b>\n\n"
            f"🆔 User ID: <code>{new_admin_id}</code>",
            parse_mode=ParseMode.HTML
        )

async def admin_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/admincikar [user_id] komutu - admin yetkisini kaldır"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece ana admin kullanabilir
    if ADMIN_USER_ID is None or update.effective_user.id != ADMIN_USER_ID:
        await update.message.reply_text("⚠️ Bu komutu sadece ana yönetici kullanabilir.")
        return
    
    # User ID kontrolü
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /admincikar [user_id]")
        return
    
    try:
        admin_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz User ID!")
        return
    
    # Kendini çıkarmaya çalışıyor mu?
    if admin_id == ADMIN_USER_ID:
        await update.message.reply_text("⚠️ Ana admin yetkisini kaldıramazsınız!")
        return
    
    # Admin çıkar
    success = db.remove_admin(admin_id)
    
    if success:
        await update.message.reply_text(
            f"✅ <b>Admin Yetkisi Kaldırıldı!</b>\n\n"
            f"🆔 User ID: <code>{admin_id}</code>\n"
            f"❌ Admin yetkileri kaldırıldı.",
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Admin yetkisi kaldırıldı: {admin_id} by {ADMIN_USER_ID}")
    else:
        await update.message.reply_text(
            f"⚠️ <b>Bu kullanıcı admin değil!</b>\n\n"
            f"🆔 User ID: <code>{admin_id}</code>",
            parse_mode=ParseMode.HTML
        )

async def admin_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/adminler komutu - tüm adminleri listele"""
    # Sadece özel mesajda çalışır
    if update.effective_chat.type != 'private':
        return
    
    # Sadece adminler kullanabilir
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    message = "👨‍💼 <b>ADMIN LİSTESİ</b>\n\n"
    
    # 1. Bot Sahibi (Ana Admin)
    message += "👑 <b>Bot Sahibi:</b>\n"
    message += f"🆔 ID: <code>{ADMIN_USER_ID}</code>\n\n"
    
    # 2. Yetkili Grup Adminleri
    group_admins = []
    if AUTHORIZED_GROUP_ID:
        try:
            admins_list = await context.bot.get_chat_administrators(AUTHORIZED_GROUP_ID)
            for admin in admins_list:
                if admin.user.id != ADMIN_USER_ID:  # Bot sahibini tekrar gösterme
                    username = f"@{admin.user.username}" if admin.user.username else admin.user.first_name
                    status = "👤 Kurucu" if admin.status == "creator" else "⚙️ Yönetici"
                    group_admins.append((admin.user.id, username, status))
        except Exception as e:
            logger.error(f"Grup adminleri alınamadı: {e}")
    
    if group_admins:
        message += f"🏢 <b>Yetkili Grup Admini:</b> ({len(group_admins)} kişi)\n\n"
        for user_id, username, status in group_admins:
            message += f"{status}\n"
            message += f"🆔 <code>{user_id}</code>\n"
            message += f"📝 {username}\n\n"
    else:
        message += "🏢 <b>Yetkili Grup Admini:</b> Yok\n\n"
    
    # 3. Veritabanı Adminleri
    db_admins = db.get_all_admins()
    if db_admins:
        message += f"💾 <b>Manuel Eklenen Adminler:</b> ({len(db_admins)} kişi)\n\n"
        for admin in db_admins:
            user_id, username, added_date, added_by = admin
            username_text = f"@{username}" if username else "Bilinmiyor"
            message += f"🔹 <code>{user_id}</code>\n"
            message += f"📝 {username_text}\n"
            message += f"📅 {added_date[:10]}\n\n"
    else:
        message += "💾 <b>Manuel Eklenen Adminler:</b> Yok\n\n"
    
    # Toplam
    total_admins = 1 + len(group_admins) + len(db_admins)
    message += f"📊 <b>Toplam Admin:</b> {total_admins} kişi\n\n"
    message += "💡 <b>Not:</b> Grup adminleri otomatik yetki alır."
    
    await update.message.reply_text(message, parse_mode=ParseMode.HTML)

async def ban_user_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/banla [user_id] [sebep] komutu - kullanıcıyı engelle"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args or len(context.args) < 1:
        await update.message.reply_text(
            "⚠️ <b>Kullanım:</b> /banla [user_id] [sebep]\n\n"
            "<b>Örnek:</b> /banla 123456789 Sahte bilgi",
            parse_mode=ParseMode.HTML
        )
        return
    
    try:
        ban_user_id = int(context.args[0])
        reason = " ".join(context.args[1:]) if len(context.args) > 1 else "Belirtilmedi"
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz User ID!")
        return
    
    if ban_user_id == update.effective_user.id:
        await update.message.reply_text("⚠️ Kendinizi engelleyemezsiniz!")
        return
    
    if is_user_admin(ban_user_id):
        await update.message.reply_text("⚠️ Başka bir admini engelleyemezsiniz!")
        return
    
    success = db.ban_user(ban_user_id, None, reason, update.effective_user.id)
    
    if success:
        await update.message.reply_text(
            f"🚫 <b>Kullanıcı Engellendi!</b>\n\n"
            f"🆔 User ID: <code>{ban_user_id}</code>\n"
            f"📝 Sebep: {reason}\n\n"
            f"Bu kullanıcı artık botu kullanamaz.",
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Kullanıcı engellendi: {ban_user_id} by {update.effective_user.id}")
        
        try:
            await context.bot.send_message(
                chat_id=ban_user_id,
                text=f"🚫 <b>Engellendiniz!</b>\n\n"
                     f"📝 Sebep: {reason}\n\n"
                     f"Bu botu kullanma yetkiniz kaldırılmıştır.",
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.warning(f"Engellenme bildirimi gönderilemedi: {e}")
    else:
        await update.message.reply_text(
            f"⚠️ <b>Bu kullanıcı zaten engellenmiş!</b>\n\n"
            f"🆔 User ID: <code>{ban_user_id}</code>",
            parse_mode=ParseMode.HTML
        )

async def unban_user_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/unban [user_id] komutu - kullanıcının engelini kaldır"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /unban [user_id]")
        return
    
    try:
        unban_user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz User ID!")
        return
    
    success = db.unban_user(unban_user_id)
    
    if success:
        await update.message.reply_text(
            f"✅ <b>Engel Kaldırıldı!</b>\n\n"
            f"🆔 User ID: <code>{unban_user_id}</code>\n\n"
            f"Kullanıcı artık botu kullanabilir.",
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Engel kaldırıldı: {unban_user_id} by {update.effective_user.id}")
        
        try:
            await context.bot.send_message(
                chat_id=unban_user_id,
                text="✅ <b>Engeliniz Kaldırıldı!</b>\n\n"
                     "Artık botu tekrar kullanabilirsiniz.\n"
                     "Başvuru yapmak için /start yazın.",
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.warning(f"Engel kaldırma bildirimi gönderilemedi: {e}")
    else:
        await update.message.reply_text(
            f"⚠️ <b>Bu kullanıcı zaten engelli değil!</b>\n\n"
            f"🆔 User ID: <code>{unban_user_id}</code>",
            parse_mode=ParseMode.HTML
        )

async def banned_users_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/banliste komutu - engellenmiş kullanıcıları göster"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    banned = db.get_banned_users()
    
    if not banned:
        await update.message.reply_text(
            "ℹ️ <b>Engellenmiş kullanıcı yok.</b>",
            parse_mode=ParseMode.HTML
        )
        return
    
    message = "🚫 <b>ENGELLENMİŞ KULLANICILAR</b>\n\n━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    for ban in banned:
        user_id, username, reason, banned_by, banned_date = ban
        username_text = f"@{username}" if username else "Bilinmiyor"
        
        message += f"🆔 <b>ID:</b> <code>{user_id}</code>\n"
        message += f"👤 <b>Username:</b> {username_text}\n"
        message += f"📝 <b>Sebep:</b> {reason}\n"
        message += f"📅 <b>Tarih:</b> {banned_date[:10]}\n"
        message += f"━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    message += f"📊 <b>Toplam:</b> {len(banned)} engellenmiş kullanıcı"
    
    await update.message.reply_text(message, parse_mode=ParseMode.HTML)

async def add_note_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/not [başvuru_id] [not] komutu - başvuruya not ekle"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args or len(context.args) < 2:
        await update.message.reply_text(
            "⚠️ <b>Kullanım:</b> /not [başvuru_id] [not içeriği]\n\n"
            "<b>Örnek:</b> /not 1 Kullanıcı ile görüşüldü",
            parse_mode=ParseMode.HTML
        )
        return
    
    try:
        application_id = int(context.args[0])
        note_text = " ".join(context.args[1:])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz başvuru ID!")
        return
    
    # Başvuru var mı kontrol et
    application = db.get_application(application_id)
    if not application:
        await update.message.reply_text(f"⚠️ {application_id} ID'li başvuru bulunamadı!")
        return
    
    # Notu ekle
    db.add_application_note(application_id, note_text, update.effective_user.id)
    
    await update.message.reply_text(
        f"✅ <b>Not Eklendi!</b>\n\n"
        f"🆔 Başvuru ID: <code>{application_id}</code>\n"
        f"📝 Not: {note_text}",
        parse_mode=ParseMode.HTML
    )
    logger.info(f"Not eklendi: Application={application_id}, by={update.effective_user.id}")

async def view_notes_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/notlar [başvuru_id] komutu - başvurunun notlarını göster"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /notlar [başvuru_id]")
        return
    
    try:
        application_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz başvuru ID!")
        return
    
    # Başvuru var mı kontrol et
    application = db.get_application(application_id)
    if not application:
        await update.message.reply_text(f"⚠️ {application_id} ID'li başvuru bulunamadı!")
        return
    
    # Notları getir
    notes = db.get_application_notes(application_id)
    
    if not notes:
        await update.message.reply_text(
            f"ℹ️ <b>Bu başvuruya henüz not eklenmemiş.</b>\n\n"
            f"🆔 Başvuru ID: <code>{application_id}</code>",
            parse_mode=ParseMode.HTML
        )
        return
    
    message = f"📝 <b>BAŞVURU NOTLARI</b>\n\n"
    message += f"🆔 Başvuru ID: <code>{application_id}</code>\n"
    message += f"━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    for note in notes:
        note_id, note_text, added_by, added_date = note
        message += f"📌 <b>Not #{note_id}</b>\n"
        message += f"📝 {note_text}\n"
        message += f"👤 Admin ID: <code>{added_by}</code>\n"
        message += f"📅 {added_date[:16]}\n"
        message += f"━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    message += f"📊 <b>Toplam:</b> {len(notes)} not"
    
    await update.message.reply_text(message, parse_mode=ParseMode.HTML)

async def bulk_approve_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/topluonayla [id1,id2,id3...] komutu - birden fazla başvuruyu onayla"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "⚠️ <b>Kullanım:</b> /topluonayla [id1,id2,id3...]\n\n"
            "<b>Örnek:</b> /topluonayla 1,2,3,4,5",
            parse_mode=ParseMode.HTML
        )
        return
    
    # ID'leri ayır
    try:
        id_string = " ".join(context.args)
        application_ids = [int(id.strip()) for id in id_string.replace(',', ' ').split()]
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz ID formatı!")
        return
    
    if len(application_ids) > 20:
        await update.message.reply_text("⚠️ Tek seferde en fazla 20 başvuru onaylayabilirsiniz!")
        return
    
    approved = []
    failed = []
    
    for app_id in application_ids:
        application = db.get_application(app_id)
        
        if not application:
            failed.append((app_id, "Başvuru bulunamadı"))
            continue
        
        if application[7] != 'pending':
            failed.append((app_id, f"Durum: {application[7]}"))
            continue
        
        # Başvuruyu onayla
        db.update_application_status(app_id, 'approved')
        approved.append(app_id)
        
        # Kullanıcıya bildirim gönder
        user_id = application[1]
        try:
            keyboard = [
                [InlineKeyboardButton("🔞 +18 Kanala İstek Gönder", url=APPROVED_GROUP_LINK)]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await context.bot.send_message(
                chat_id=user_id,
                text="🎉✨ <b>TEBRİKLER KİRVEM!</b> ✨🎉\n\n"
                     "🎊 Artık aramızdasınız! Hoş geldiniz!\n\n"
                     "🔞 +18 Özel Kanalımıza Katılın\n"
                     "💬 Sohbet Grubumuzla Tanışın\n\n"
                     "👆 İki butona da tıklayabilirsiniz!\n\n"
                     "💫 Eğlenceli sohbetler sizi bekliyor!",
                reply_markup=reply_markup,
                parse_mode=ParseMode.HTML
            )
        except:
            pass
    
    # Sonuç mesajı
    result_msg = f"📊 <b>TOPLU ONAY SONUCU</b>\n\n"
    result_msg += f"✅ Onaylanan: {len(approved)} başvuru\n"
    result_msg += f"❌ Başarısız: {len(failed)} başvuru\n\n"
    
    if approved:
        result_msg += f"✅ <b>Onaylananlar:</b> {', '.join(map(str, approved))}\n\n"
    
    if failed:
        result_msg += "❌ <b>Başarısızlar:</b>\n"
        for fail_id, reason in failed:
            result_msg += f"  • ID {fail_id}: {reason}\n"
    
    await update.message.reply_text(result_msg, parse_mode=ParseMode.HTML)
    logger.info(f"Toplu onay: {len(approved)} başarılı, {len(failed)} başarısız by {update.effective_user.id}")

async def bulk_reject_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/toplureddet [id1,id2,id3...] komutu - birden fazla başvuruyu reddet"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "⚠️ <b>Kullanım:</b> /toplureddet [id1,id2,id3...]\n\n"
            "<b>Örnek:</b> /toplureddet 1,2,3,4,5",
            parse_mode=ParseMode.HTML
        )
        return
    
    # ID'leri ayır
    try:
        id_string = " ".join(context.args)
        application_ids = [int(id.strip()) for id in id_string.replace(',', ' ').split()]
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz ID formatı!")
        return
    
    if len(application_ids) > 20:
        await update.message.reply_text("⚠️ Tek seferde en fazla 20 başvuru reddedebilirsiniz!")
        return
    
    rejected = []
    failed = []
    
    for app_id in application_ids:
        application = db.get_application(app_id)
        
        if not application:
            failed.append((app_id, "Başvuru bulunamadı"))
            continue
        
        if application[7] != 'pending':
            failed.append((app_id, f"Durum: {application[7]}"))
            continue
        
        # Başvuruyu reddet
        db.update_application_status(app_id, 'rejected', 'Toplu red')
        rejected.append(app_id)
        
        # Kullanıcıya bildirim gönder
        user_id = application[1]
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"❌ <b>Başvurunuz Reddedildi</b>\n\n"
                     f"📝 Sebep: Toplu işlem\n\n"
                     f"⏱️ <b>20 dakika sonra</b> tekrar başvuru yapabilirsiniz.\n\n"
                     f"Tekrar başvurmak için /start yazın.",
                parse_mode=ParseMode.HTML
            )
        except:
            pass
    
    # Sonuç mesajı
    result_msg = f"📊 <b>TOPLU RED SONUCU</b>\n\n"
    result_msg += f"❌ Reddedilen: {len(rejected)} başvuru\n"
    result_msg += f"⚠️ Başarısız: {len(failed)} başvuru\n\n"
    
    if rejected:
        result_msg += f"❌ <b>Redded Edilen:</b> {', '.join(map(str, rejected))}\n\n"
    
    if failed:
        result_msg += "⚠️ <b>Başarısızlar:</b>\n"
        for fail_id, reason in failed:
            result_msg += f"  • ID {fail_id}: {reason}\n"
    
    await update.message.reply_text(result_msg, parse_mode=ParseMode.HTML)
    logger.info(f"Toplu red: {len(rejected)} başarılı, {len(failed)} başarısız by {update.effective_user.id}")

async def user_profile_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/profil [user_id] komutu - kullanıcı profil ve geçmişi"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    if not context.args:
        await update.message.reply_text("⚠️ Kullanım: /profil [user_id]")
        return
    
    try:
        target_user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Geçersiz User ID!")
        return
    
    # Kullanıcının tüm başvurularını getir
    conn = db.sqlite3.connect(db.DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, username, site_name, site_username, status, application_date, approved_date
        FROM applications
        WHERE user_id = ?
        ORDER BY application_date DESC
    ''', (target_user_id,))
    
    applications = cursor.fetchall()
    conn.close()
    
    if not applications:
        await update.message.reply_text(
            f"ℹ️ <b>Başvuru Bulunamadı</b>\n\n"
            f"🆔 User ID: <code>{target_user_id}</code>\n\n"
            f"Bu kullanıcının hiç başvurusu yok.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # İstatistikler
    total = len(applications)
    pending = sum(1 for app in applications if app[4] == 'pending')
    approved = sum(1 for app in applications if app[4] == 'approved')
    rejected = sum(1 for app in applications if app[4] == 'rejected')
    
    # Engel kontrolü
    is_banned = db.is_user_banned(target_user_id)
    
    # Mesaj oluştur
    username = applications[0][1] if applications[0][1] else "Bilinmiyor"
    
    message = f"👤 <b>KULLANICI PROFİLİ</b>\n\n"
    message += f"🆔 <b>User ID:</b> <code>{target_user_id}</code>\n"
    message += f"📝 <b>Username:</b> {username}\n"
    message += f"🚫 <b>Durum:</b> {'⛔️ Engellenmiş' if is_banned else '✅ Aktif'}\n\n"
    message += f"━━━━━━━━━━━━━━━━━━━━━\n\n"
    message += f"📊 <b>BAŞVURU İSTATİSTİKLERİ</b>\n\n"
    message += f"📌 Toplam Başvuru: {total}\n"
    message += f"⏳ Bekleyen: {pending}\n"
    message += f"✅ Onaylanan: {approved}\n"
    message += f"❌ Reddedilen: {rejected}\n\n"
    message += f"━━━━━━━━━━━━━━━━━━━━━\n\n"
    message += f"📋 <b>BAŞVURU GEÇMİŞİ</b>\n\n"
    
    for app in applications[:5]:  # Son 5 başvuru
        app_id, username, site_name, site_username, status, app_date, appr_date = app
        
        status_emoji = {"pending": "⏳", "approved": "✅", "rejected": "❌"}
        status_text = {"pending": "Bekliyor", "approved": "Onaylandı", "rejected": "Reddedildi"}
        
        message += f"🔹 <b>ID {app_id}</b> - {status_emoji.get(status, '❓')} {status_text.get(status, status)}\n"
        message += f"   🏢 Site: {db.SITES.get(site_name, {}).get('name', site_name) if hasattr(db, 'SITES') else site_name}\n"
        message += f"   📅 Tarih: {app_date[:10]}\n"
        if appr_date and status == 'approved':
            message += f"   ✅ Onay: {appr_date[:10]}\n"
        message += f"\n"
    
    if total > 5:
        message += f"<i>... ve {total - 5} başvuru daha</i>\n\n"
    
    await update.message.reply_text(message, parse_mode=ParseMode.HTML)
    logger.info(f"Profil görüntülendi: User={target_user_id} by {update.effective_user.id}")

async def yaziyaz_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/yaziyaz komutu - Tüm kayıtlı gruplara mesaj gönder (Rose bot gibi)"""
    if update.effective_chat.type != 'private':
        return
    
    if not await is_user_admin(update.effective_user.id, context):
        await update.message.reply_text("⚠️ Bu komutu sadece adminler kullanabilir.")
        return
    
    # Komut formatı: /yaziyaz Mesaj metni
    if not context.args:
        await update.message.reply_text(
            "📝 <b>Kullanım:</b>\n"
            "<code>/yaziyaz Mesaj metni</code>\n\n"
            "💡 <b>Örnekler:</b>\n"
            "• <code>/yaziyaz Merhaba! 🎉</code>\n"
            "• <code>/yaziyaz [Emoji](Yazı)(Link)</code>\n"
            "• <code>/yaziyaz <b>Kalın</b> <i>İtalik</i></code>\n\n"
            "📌 <b>Desteklenen Formatlar:</b>\n"
            "• HTML: &lt;b&gt;kalın&lt;/b&gt;, &lt;i&gt;italik&lt;/i&gt;, &lt;a href='link'&gt;metin&lt;/a&gt;\n"
            "• Markdown: **kalın**, *italik*, [metin](link)\n"
            "• Emoji: 🎉 ✅ 🔥 💎\n\n"
            "⚠️ <b>Not:</b> Mesaj tüm kayıtlı gruplara gönderilecektir!",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Mesaj metnini al
    message_text = ' '.join(context.args)
    
    # Tüm kayıtlı grupları al
    groups = db.get_all_groups()
    
    if not groups:
        await update.message.reply_text(
            "⚠️ <b>Kayıtlı Grup Yok!</b>\n\n"
            "Botu bir gruba ekle ve admin yap. Grup otomatik olarak kaydedilecektir.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Önizleme göster
    preview_msg = await update.message.reply_text(
        f"📤 <b>Mesaj Gönderiliyor...</b>\n\n"
        f"📋 <b>Önizleme:</b>\n"
        f"{message_text}\n\n"
        f"📊 <b>Hedef:</b> {len(groups)} grup\n\n"
        f"⏳ Lütfen bekleyin...",
        parse_mode=ParseMode.HTML
    )
    
    # Her gruba mesaj gönder
    success_count = 0
    fail_count = 0
    invalid_groups = []
    results = []
    
    for group_id, group_name in groups:
        try:
            # HTML formatında gönder
            await context.bot.send_message(
                chat_id=group_id,
                text=message_text,
                parse_mode=ParseMode.HTML,
                disable_web_page_preview=False
            )
            success_count += 1
            results.append(f"✅ {group_name} ({group_id})")
            logger.info(f"Mesaj gönderildi: {group_name} ({group_id})")
            
            # Rate limiting (Telegram API limit)
            await asyncio.sleep(0.1)
            
        except Exception as e:
            fail_count += 1
            error_msg = str(e).lower()
            
            # Geçersiz grup kontrolü
            if "chat not found" in error_msg or "bot was kicked" in error_msg or "bot is not a member" in error_msg:
                invalid_groups.append((group_id, group_name))
                results.append(f"❌ {group_name} ({group_id}) - Geçersiz grup")
                logger.warning(f"Geçersiz grup: {group_name} ({group_id})")
            else:
                results.append(f"⚠️ {group_name} ({group_id}) - {str(e)[:50]}")
                logger.error(f"Mesaj gönderme hatası ({group_name}): {e}")
    
    # Geçersiz grupları veritabanından sil
    if invalid_groups:
        for group_id, group_name in invalid_groups:
            db.remove_group(group_id)
            logger.info(f"Geçersiz grup silindi: {group_name} ({group_id})")
    
    # Sonuç mesajı oluştur
    result_message = f"📊 <b>GÖNDERİM SONUÇLARI</b>\n\n"
    result_message += f"✅ <b>Başarılı:</b> {success_count}\n"
    result_message += f"❌ <b>Başarısız:</b> {fail_count}\n"
    
    if invalid_groups:
        result_message += f"🗑️ <b>Silinen Gruplar:</b> {len(invalid_groups)}\n"
    
    result_message += f"\n━━━━━━━━━━━━━━━━━━━━━\n\n"
    
    # İlk 10 sonucu göster
    if results:
        result_message += f"<b>Detaylar:</b>\n"
        for result in results[:10]:
            result_message += f"{result}\n"
        
        if len(results) > 10:
            result_message += f"\n<i>... ve {len(results) - 10} sonuç daha</i>"
    
    # Önizleme mesajını güncelle
    try:
        await preview_msg.edit_text(result_message, parse_mode=ParseMode.HTML)
    except:
        await update.message.reply_text(result_message, parse_mode=ParseMode.HTML)
    
    logger.info(f"Toplu mesaj gönderimi tamamlandı: {success_count} başarılı, {fail_count} başarısız")

def main():
    """Ana fonksiyon"""
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Graceful shutdown handler
    def signal_handler(signum, frame):
        logger.info(f"Signal {signum} received. Shutting down gracefully...")
        application.stop_running()
        sys.exit(0)
    
    # Signal handler'ları kaydet
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    # Komut handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("yardim", help_command))
    application.add_handler(CommandHandler("help", help_command))  # İngilizce de çalışsın
    application.add_handler(CommandHandler("adminyardim", admin_help_command))
    application.add_handler(CommandHandler("benkim", benkim))
    application.add_handler(CommandHandler("grupid", grupid))
    application.add_handler(CommandHandler("testfoto", test_welcome_photos))
    application.add_handler(CommandHandler("durumum", user_application_status))
    application.add_handler(CommandHandler("sifirla", reset_my_application))
    application.add_handler(CommandHandler("basvurusil", admin_reset_application))
    application.add_handler(CommandHandler("bekleyenler", admin_pending_applications))
    application.add_handler(CommandHandler("ss", admin_view_screenshots))
    application.add_handler(CommandHandler("incele", admin_view_screenshots))
    application.add_handler(CommandHandler("onayla", admin_approve))
    application.add_handler(CommandHandler("reddet", admin_reject))
    application.add_handler(CommandHandler("gecmis", admin_history))
    application.add_handler(CommandHandler("istatistikler", admin_statistics))
    application.add_handler(CommandHandler("exportexcel", admin_export_excel))
    application.add_handler(CommandHandler("kirvebasla", admin_send_welcome))
    application.add_handler(CommandHandler("kirvebasla2", admin_send_welcome_theme5))  # Çoklu başvuru sistemi
    application.add_handler(CommandHandler("kirvebasla3", admin_send_welcome_theme2))
    application.add_handler(CommandHandler("kirvebasla4", admin_send_welcome_theme3))
    application.add_handler(CommandHandler("kirvebasla5", admin_send_welcome_theme4))
    application.add_handler(CommandHandler("kirvebaslat", admin_send_welcome_in_group))
    application.add_handler(CommandHandler("adminekle", admin_add))
    application.add_handler(CommandHandler("admincikar", admin_remove))
    application.add_handler(CommandHandler("adminler", admin_list))
    application.add_handler(CommandHandler("banla", ban_user_command))
    application.add_handler(CommandHandler("unban", unban_user_command))
    application.add_handler(CommandHandler("banliste", banned_users_list))
    application.add_handler(CommandHandler("not", add_note_command))
    application.add_handler(CommandHandler("notlar", view_notes_command))
    application.add_handler(CommandHandler("topluonayla", bulk_approve_command))
    application.add_handler(CommandHandler("toplureddet", bulk_reject_command))
    application.add_handler(CommandHandler("profil", user_profile_command))
    application.add_handler(CommandHandler("yaziyaz", yaziyaz_command))
    
    # Özel komutlar (! ile başlayanlar)
    from telegram.ext import MessageHandler, filters
    application.add_handler(MessageHandler(
        filters.Regex(r'^!cc(\s+\d+)?$') & filters.ChatType.GROUPS,
        admin_clear_messages
    ))
    
    # Callback query handler
    application.add_handler(CallbackQueryHandler(button_callback))
    
    # Servis mesajlarını sil (join/leave) - EN ÖNCE
    application.add_handler(MessageHandler(
        (filters.StatusUpdate.NEW_CHAT_MEMBERS | filters.StatusUpdate.LEFT_CHAT_MEMBER),
        delete_service_messages
    ))
    
    # Mesaj handlers
    application.add_handler(MessageHandler(filters.PHOTO & filters.ChatType.PRIVATE, handle_photo))
    application.add_handler(MessageHandler(filters.TEXT & filters.ChatType.PRIVATE & ~filters.COMMAND, handle_text))
    
    # Bot gruba eklendiğinde
    application.add_handler(ChatMemberHandler(my_chat_member_handler, ChatMemberHandler.MY_CHAT_MEMBER))
    
    logger.info("Bot başlatılıyor...")
    
    # Adminlere başlangıç bildirimi gönder
    async def send_startup_notification():
        startup_message = (
            f"🤖 <b>KirveKing Bot Yeniden Başlatıldı!</b>\n\n"
            f"📊 <b>Bot Bilgileri:</b>\n"
            f"• Sürüm: {BOT_VERSION}\n"
            f"• Son Güncelleme: {LAST_UPDATE}\n"
            f"• Durum: ✅ Aktif\n\n"
            f"🔧 <b>Yeni Özellikler:</b>\n"
            f"• Genel sohbet desteği\n"
            f"• Akıllı yardım sistemi\n"
            f"• Database güvenlik kilidi\n"
            f"• Geliştirilmiş hata yönetimi\n\n"
            f"💫 <b>Bot hazır ve çalışıyor!</b>"
        )
        
        # Tüm adminlere bildirim gönder
        admin_ids = db.get_all_admin_ids()
        for admin_id in admin_ids:
            try:
                await application.bot.send_message(
                    chat_id=admin_id,
                    text=startup_message,
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                logger.error(f"Startup notification failed for admin {admin_id}: {e}")
    
    # Bot başlat
    logger.info("Bot başlatılıyor...")
    
    # Başlangıç bildirimi her başlatmada gönder
    if True:
        async def send_startup_message():
            try:
                startup_message = (
                    f"🤖 <b>KirveKing Bot Başlatıldı!</b>\n\n"
                    f"📊 <b>Bot Bilgileri:</b>\n"
                    f"• Sürüm: {BOT_VERSION}\n"
                    f"• Son Güncelleme: {LAST_UPDATE}\n"
                    f"• Durum: ✅ Aktif\n\n"
                    f"🔧 <b>Özellikler:</b>\n"
                    f"• Genel sohbet desteği\n"
                    f"• Akıllı yardım sistemi\n"
                    f"• Database güvenlik kilidi\n"
                    f"• Geliştirilmiş hata yönetimi\n\n"
                    f"💫 <b>Bot hazır ve çalışıyor!</b>"
                )
                
                # Ana admin'e önce gönder
                try:
                    await application.bot.send_message(
                        chat_id=ADMIN_USER_ID,
                        text=startup_message,
                        parse_mode=ParseMode.HTML
                    )
                    logger.info(f"Startup notification sent to main admin: {ADMIN_USER_ID}")
                except Exception as e:
                    logger.error(f"Startup notification failed for main admin {ADMIN_USER_ID}: {e}")
                
                # Diğer adminlere de gönder
                try:
                    admin_ids = db.get_all_admin_ids()
                    for admin_id in admin_ids:
                        if admin_id != ADMIN_USER_ID:  # Ana admin'i tekrar gönderme
                            try:
                                await application.bot.send_message(
                                    chat_id=admin_id,
                                    text=startup_message,
                                    parse_mode=ParseMode.HTML
                                )
                            except Exception as e:
                                logger.error(f"Startup notification failed for admin {admin_id}: {e}")
                except Exception as e:
                    logger.error(f"Error getting admin IDs: {e}")
                
                # Artık dosya oluşturmuyoruz (her seferinde gönder)
                    
            except Exception as e:
                logger.error(f"Startup notification error: {e}")
        
        # Bot başladıktan sonra bildirim gönder (post_init hook kullan)
        async def post_init(application):
            await asyncio.sleep(5)  # 5 saniye bekle
            await send_startup_message()
        
        application.post_init = post_init
    
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
